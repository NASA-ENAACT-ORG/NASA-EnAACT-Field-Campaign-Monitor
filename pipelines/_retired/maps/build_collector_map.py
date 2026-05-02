#!/usr/bin/env python3
"""
build_collector_map.py
Generates collector_map.html — a standalone split-screen Leaflet viewer
with collector pins (walk count inside) and a slide-in detail sidebar.

Run:  python build_collector_map.py
Out:  collector_map.html  (same directory)
"""

import json, hashlib, math, xml.etree.ElementTree as ET, sys
import pandas as pd
from datetime import date, timedelta
from pathlib import Path

BASE = Path(__file__).parent

# Add repo root to sys.path so shared package is importable
_REPO_ROOT = Path(__file__).resolve().parent.parent.parent
if str(_REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(_REPO_ROOT))

from shared.paths import (
    COLLECTOR_MAP_HTML as OUT,
    ROUTES_DATA_JSON, ROUTES_KML_DIR,
    WALKS_LOG, SCHEDULE_OUTPUT_JSON,
    V2_PREFERRED_ROUTES, PREFERRED_ROUTES,
    AVAILABILITY_XLSX,
)
from shared.gcs import pull_if_available as gcs_pull
from shared.registry import (
    ACTIVE_COLLECTORS,
    COLLECTOR_DISPLAY_NAMES,
    COLLECTOR_KML_NAME_TO_ID,
    COLLECTOR_PIN_COLORS as C_COLOR,
)

# Pull the latest bucket copies so the map reflects live schedule + walk state.
gcs_pull("Walks_Log.txt",        WALKS_LOG)
gcs_pull("schedule_output.json", SCHEDULE_OUTPUT_JSON)

# ── Collector registry ─────────────────────────────────────────────────────────
# Backpack A = CCNY (purple pins), Backpack B = LaGCC (red pins)
# ANG is CCNY last-resort staff (purple). PRA/NAT/NRS are professors and are not scheduled.
COLLECTORS = {cid: COLLECTOR_DISPLAY_NAMES.get(cid, cid) for cid in ACTIVE_COLLECTORS}

DAYS = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
TODS = ["AM", "MD", "PM"]

# ── Load routes ────────────────────────────────────────────────────────────────
with open(ROUTES_DATA_JSON) as f:
    routes_raw = json.load(f)

def centroid(lines):
    lats, lngs = [], []
    for seg in lines:
        for lat, lng in seg:
            lats.append(lat); lngs.append(lng)
    if not lats:
        return (40.7128, -74.006)
    return (sum(lats) / len(lats), sum(lngs) / len(lngs))

route_centroid = {k: centroid(v["lines"]) for k, v in routes_raw.items()}
routes_meta    = {k: {"boro": v["boro"], "name": v["name"]} for k, v in routes_raw.items()}

# Neighborhood-code → full route key  (e.g. "BS" → "BK_BS")
neigh_to_route = {k.split("_")[1]: k for k in routes_raw}

# ── Load real collector home locations from KML ────────────────────────────────
kml_homes: dict[str, tuple[float, float]] = {}   # cid → (lat, lng)
_kml_path = ROUTES_KML_DIR / "Collector_Locs.kml"
if _kml_path.exists():
    _ns = {"k": "http://www.opengis.net/kml/2.2"}
    for _pm in ET.parse(_kml_path).findall(".//k:Placemark", _ns):
        _nm  = (_pm.findtext("k:name", "", _ns) or "").strip()
        _crd = (_pm.findtext(".//k:coordinates", "", _ns) or "").strip()
        _cid = COLLECTOR_KML_NAME_TO_ID.get(_nm)
        if _cid and _crd:
            _lng, _lat = float(_crd.split(",")[0]), float(_crd.split(",")[1])
            kml_homes[_cid] = (round(_lat, 6), round(_lng, 6))

# ── Parse Walks_Log.txt ────────────────────────────────────────────────────────
walks = []
log_path = WALKS_LOG
if log_path.exists():
    for ln in log_path.read_text().splitlines():
        ln = ln.strip()
        if not ln or ln.startswith("RECAL"):
            continue
        p = ln.split("_")
        if len(p) >= 6 and p[0] == "X":
            walks.append({
                "collector": p[1],
                "route":     f"{p[2]}_{p[3]}",
                "boro":      p[2],
                "neigh":     p[3],
                "date":      f"{p[4][:4]}-{p[4][4:6]}-{p[4][6:]}",
                "tod":       p[5],
            })

# ── Load schedule ─────────────────────────────────────────────────────────────
sched_data = {}
sp = SCHEDULE_OUTPUT_JSON
if sp.exists():
    sched_data = json.load(open(sp))
sched_assignments = sched_data.get("assignments", [])

# ── Load V2 preferred routes (0-3), fall back to V1 (binary→1) ───────────────
# pref[cid][route_key] = 0..3
pref: dict[str, dict[str, int]] = {cid: {} for cid in COLLECTORS}
try:
    v2_path = V2_PREFERRED_ROUTES
    if v2_path.exists():
        df2 = pd.read_excel(v2_path, sheet_name="Route Ratings", header=0, index_col=0)
        for cid in df2.index:
            cid = str(cid).strip()
            if cid not in COLLECTORS:
                continue
            for neigh_code, val in df2.loc[cid].items():
                neigh_code = str(neigh_code).strip()
                route_key  = neigh_to_route.get(neigh_code)
                if route_key and pd.notna(val):
                    try:
                        pref[cid][route_key] = int(float(val))
                    except (ValueError, TypeError):
                        pass
except Exception as e:
    print(f"[warn] V2 pref load failed: {e}")

try:
    v1_path = PREFERRED_ROUTES
    if v1_path.exists():
        df1 = pd.read_excel(v1_path, header=None)
        # Locate header row: first row that contains a collector ID
        col_to_cid: dict[int, str] = {}
        route_col = -1
        data_start = -1
        for i, row in df1.iterrows():
            found = {j: str(v).strip() for j, v in enumerate(row)
                     if str(v).strip() in COLLECTORS}
            if found:
                col_to_cid = found
                data_start = int(i) + 1
                # Route code is in col 2
                route_col = 2
                break
        if col_to_cid and data_start >= 0:
            for i in range(data_start, len(df1)):
                row = df1.iloc[i]
                neigh_code = str(row.iloc[route_col]).strip() if route_col < len(row) else ""
                route_key  = neigh_to_route.get(neigh_code)
                if not route_key:
                    continue
                for col_j, cid in col_to_cid.items():
                    # Only fill in for collectors not already in V2
                    if pref[cid].get(route_key) is not None:
                        continue
                    val = row.iloc[col_j] if col_j < len(row) else None
                    if pd.notna(val) and str(val).strip() == "1":
                        pref[cid][route_key] = 1
except Exception as e:
    print(f"[warn] V1 pref load failed: {e}")

# ── Load availability ──────────────────────────────────────────────────────────
avail: dict[str, dict[str, list[bool]]] = {}
try:
    awb_path = AVAILABILITY_XLSX
    if awb_path.exists():
        import openpyxl
        awb = openpyxl.load_workbook(awb_path, data_only=True)
        for sn in awb.sheetnames:
            cid = sn.strip().upper()
            ws  = awb[sn]
            cav: dict[str, list[bool]] = {}
            for ri, tod in enumerate(TODS, start=2):
                cav[tod] = [
                    ws.cell(row=ri, column=ci).value == 1
                    for ci in range(2, 9)
                ]
            avail[cid] = cav
except Exception as e:
    print(f"[warn] Availability load failed: {e}")

# ── Compute per-collector pin position ────────────────────────────────────────
def pin_pos(cid: str) -> tuple[float, float]:
    # Prefer real KML home location
    if cid in kml_homes:
        return kml_homes[cid]
    # Fallback: centroid of walked routes, then preferred routes, then NYC center
    cw = [w for w in walks if w["collector"] == cid]
    pts = [route_centroid[w["route"]] for w in cw if w["route"] in route_centroid]
    if not pts:
        p_routes = sorted(pref.get(cid, {}).items(), key=lambda x: -x[1])
        for rk, score in p_routes:
            if rk in route_centroid:
                pts.append(route_centroid[rk])
                if len(pts) >= 3:
                    break
    if pts:
        lat = sum(p[0] for p in pts) / len(pts)
        lng = sum(p[1] for p in pts) / len(pts)
    else:
        lat, lng = 40.7128, -74.006
    h   = int(hashlib.md5(cid.encode()).hexdigest()[:4], 16)
    lat += ((h % 100) - 50) * 0.0011
    lng += ((h // 100 % 100) - 50) * 0.0016
    return round(lat, 6), round(lng, 6)

# ── Build collector data payload ───────────────────────────────────────────────
today      = date.today()
week_start = today - timedelta(days=today.weekday())

collector_payload: dict[str, dict] = {}
for cid, name in COLLECTORS.items():
    cw         = [w for w in walks if w["collector"] == cid]
    lat, lng   = pin_pos(cid)
    this_month = [w for w in cw if w["date"][:7] == today.strftime("%Y-%m")]
    this_week  = [w for w in cw if date.fromisoformat(w["date"]) >= week_start]

    tod_counts = {t: sum(1 for w in cw if w["tod"] == t) for t in TODS}

    route_counts: dict[str, int] = {}
    for w in cw:
        route_counts[w["route"]] = route_counts.get(w["route"], 0) + 1

    upcoming = sorted(
        [a for a in sched_assignments if a.get("collector") == cid],
        key=lambda a: a.get("date", "")
    )[:5]

    # Preference list: sorted descending by score, only score ≥ 1
    pref_list = [
        {"route": rk, "score": sc, "name": routes_meta.get(rk, {}).get("name", rk)}
        for rk, sc in sorted(pref.get(cid, {}).items(), key=lambda x: -x[1])
        if sc >= 1
    ]

    collector_payload[cid] = {
        "id":       cid,
        "name":     name,
        "color":    C_COLOR[cid],
        "lat":      lat,
        "lng":      lng,
        "total":    len(cw),
        "month":    len(this_month),
        "week":     len(this_week),
        "tod":      tod_counts,
        "recent":   sorted(cw, key=lambda w: w["date"], reverse=True)[:6],
        "upcoming": upcoming,
        "routes":   dict(sorted(route_counts.items(), key=lambda x: -x[1])),
        "avail":    avail.get(cid, {}),
        "pref":     pref_list,
    }

# ── Emit HTML ──────────────────────────────────────────────────────────────────
CDATA = json.dumps(collector_payload)
RMETA = json.dumps(routes_meta)

HTML = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8"/>
<meta name="viewport" content="width=device-width,initial-scale=1"/>
<title>NASA EnAACT – Collector Map</title>
<link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css"/>
<script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"></script>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{background:#0f1117;color:#e2e8f0;font-family:-apple-system,BlinkMacSystemFont,'Inter','Segoe UI',sans-serif;height:100vh;display:flex;flex-direction:column;overflow:hidden}}

/* ── Header ── */
#app-header{{
  background:#13151f;border-bottom:1px solid #2a2d3e;
  padding:10px 20px;display:flex;align-items:center;gap:12px;flex-shrink:0;
  z-index:2000;
}}
#app-header .logo{{font-size:11px;font-weight:700;letter-spacing:.08em;color:#4f8ef7;text-transform:uppercase}}
#app-header .title{{font-size:13px;font-weight:600;color:#c8cfe8}}
#app-header .sep{{color:#3a3d52;font-size:16px}}
#app-header .subtitle{{font-size:12px;color:#6b7099}}

/* ── Layout ── */
#stage{{flex:1;position:relative;overflow:hidden}}

/* ── Map ── */
#map{{position:absolute;inset:0;background:#0d0f18}}

/* ── Sidebar ── */
#sidebar{{
  position:absolute;top:0;right:0;width:360px;height:100%;
  background:#13151f;border-left:1px solid #2a2d3e;
  transform:translateX(100%);transition:transform .28s cubic-bezier(.4,0,.2,1);
  z-index:1000;display:flex;flex-direction:column;overflow:hidden;
}}
#sidebar.open{{transform:translateX(0)}}

#sb-close{{
  position:absolute;top:10px;right:12px;
  width:28px;height:28px;border-radius:50%;
  background:#1e2135;border:1px solid #2a2d3e;
  color:#8891b4;cursor:pointer;font-size:14px;
  display:flex;align-items:center;justify-content:center;
  transition:background .15s,color .15s;z-index:10;
}}
#sb-close:hover{{background:#2a2d3e;color:#e2e8f0}}

#sb-scroll{{flex:1;overflow-y:auto;padding-bottom:24px}}
#sb-scroll::-webkit-scrollbar{{width:4px}}
#sb-scroll::-webkit-scrollbar-track{{background:transparent}}
#sb-scroll::-webkit-scrollbar-thumb{{background:#2a2d3e;border-radius:2px}}

/* ── Sidebar header ── */
.sb-head{{
  padding:20px 20px 16px;
  border-bottom:1px solid #1e2135;
  display:flex;align-items:center;gap:14px;
  border-left:4px solid var(--c);
}}
.sb-avatar{{
  width:46px;height:46px;border-radius:50%;
  background:var(--c);display:flex;align-items:center;justify-content:center;
  font-weight:800;font-size:14px;color:#fff;flex-shrink:0;
  box-shadow:0 0 0 3px rgba(255,255,255,.1);
}}
.sb-name{{font-size:15px;font-weight:700;color:#e2e8f0}}
.sb-id{{font-size:11px;color:#6b7099;margin-top:2px;font-family:monospace;letter-spacing:.06em}}

/* ── Stat cards ── */
.stats-row{{display:flex;gap:8px;padding:14px 16px}}
.stat-card{{
  flex:1;background:#1a1d2e;border:1px solid #2a2d3e;border-radius:8px;
  padding:10px 8px;text-align:center;
}}
.stat-val{{font-size:22px;font-weight:800;color:#e2e8f0;line-height:1}}
.stat-lbl{{font-size:10px;color:#6b7099;margin-top:4px;text-transform:uppercase;letter-spacing:.06em}}

/* ── Sections ── */
.section{{padding:12px 16px 4px}}
.sec-title{{
  font-size:10px;font-weight:700;letter-spacing:.1em;
  text-transform:uppercase;color:#4f8ef7;margin-bottom:10px;
}}

/* ── TOD bars ── */
.tod-bar-row{{display:flex;align-items:center;gap:8px;margin-bottom:7px}}
.tod-lbl{{width:24px;font-size:11px;font-weight:600;color:#8891b4;flex-shrink:0}}
.tod-bg{{flex:1;height:8px;background:#1e2135;border-radius:4px;overflow:hidden}}
.tod-fill{{height:100%;border-radius:4px;transition:width .4s ease}}
.tod-num{{width:20px;font-size:11px;color:#8891b4;text-align:right;flex-shrink:0}}

/* ── Availability grid ── */
.avail-grid{{display:grid;grid-template-columns:32px repeat(7,1fr);gap:3px;font-size:10px}}
.ag-head{{color:#6b7099;text-align:center;font-weight:600;padding:2px 0}}
.ag-tod{{color:#8891b4;font-weight:700;display:flex;align-items:center;font-size:10px}}
.ag-cell{{
  height:22px;border-radius:4px;display:flex;align-items:center;justify-content:center;
  font-size:11px;font-weight:600;
}}
.ag-yes{{background:#1a3a2a;color:#34d399}}
.ag-no{{background:#1e2135;color:#3a3d52}}

/* ── Walk rows ── */
.walk-row{{
  display:flex;align-items:center;gap:8px;padding:6px 0;
  border-bottom:1px solid #1a1d2e;
}}
.walk-row:last-child{{border-bottom:none}}
.tod-badge{{
  font-size:9px;font-weight:800;padding:2px 5px;border-radius:4px;
  text-transform:uppercase;flex-shrink:0;letter-spacing:.04em;
}}
.tod-AM{{background:#332a00;color:#fbbf24}}
.tod-MD{{background:#0d1f3f;color:#60a5fa}}
.tod-PM{{background:#1e1035;color:#a78bfa}}
.walk-route{{flex:1;font-size:12px;color:#c8cfe8;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}}
.walk-date{{font-size:10px;color:#6b7099;flex-shrink:0;font-family:monospace}}
.walk-bp{{font-size:9px;color:#4f8ef7;background:#0d1f3f;padding:1px 5px;border-radius:3px;flex-shrink:0}}

/* ── Preference list ── */
.pref-row{{display:flex;align-items:center;gap:8px;padding:5px 0;border-bottom:1px solid #1a1d2e}}
.pref-row:last-child{{border-bottom:none}}
.pref-name{{flex:1;font-size:12px;color:#c8cfe8}}
.pref-score{{display:flex;gap:3px}}
.pref-dot{{width:8px;height:8px;border-radius:50%}}
.dot-on{{background:#4f8ef7}}
.dot-off{{background:#2a2d3e}}

/* ── Legend ── */
#legend{{
  position:absolute;bottom:24px;left:12px;
  background:rgba(19,21,31,.92);border:1px solid #2a2d3e;
  border-radius:10px;padding:10px 12px;z-index:500;
  backdrop-filter:blur(6px);min-width:130px;
}}
.leg-title{{font-size:9px;font-weight:700;text-transform:uppercase;letter-spacing:.1em;color:#4f8ef7;margin-bottom:8px}}
.leg-row{{display:flex;align-items:center;gap:7px;margin-bottom:5px;cursor:pointer;opacity:.85;transition:opacity .15s}}
.leg-row:hover{{opacity:1}}
.leg-dot{{width:10px;height:10px;border-radius:50%;flex-shrink:0}}
.leg-name{{font-size:11px;color:#c8cfe8}}
.leg-count{{font-size:10px;color:#6b7099;margin-left:auto;padding-left:6px}}

/* ── Map pin styling ── */
.cpin-icon{{background:transparent!important;border:none!important}}
</style>
</head>
<body>

<div id="app-header">
  <span class="logo">NASA EnAACT</span>
  <span class="sep">|</span>
  <span class="title">Field Campaign Monitor</span>
  <span class="sep">—</span>
  <span class="subtitle">Collector Map</span>
</div>

<div id="stage">
  <div id="map"></div>

  <div id="sidebar">
    <button id="sb-close" title="Close">✕</button>
    <div id="sb-scroll">
      <div id="sb-content"></div>
    </div>
  </div>

  <div id="legend">
    <div class="leg-title">Collectors</div>
    <div id="leg-body"></div>
  </div>
</div>

<script>
const COLLECTORS = {CDATA};
const ROUTES_META = {RMETA};

// ── Helpers ────────────────────────────────────────────────────────────────────
const $ = id => document.getElementById(id);
const TOD_COLORS = {{AM:'#fbbf24', MD:'#60a5fa', PM:'#a78bfa'}};

function fmtDate(iso){{
  if(!iso) return '';
  const [y,m,d] = iso.split('-');
  return `${{m}}/${{d}}/${{y.slice(2)}}`;
}}

// ── Map init ───────────────────────────────────────────────────────────────────
const map = L.map('map', {{
  center: [40.720, -73.980],
  zoom: 11,
  zoomControl: false,
  attributionControl: false,
}});
L.tileLayer('https://{{s}}.basemaps.cartocdn.com/dark_all/{{z}}/{{x}}/{{y}}{{r}}.png', {{
  maxZoom: 18, subdomains: 'abcd'
}}).addTo(map);
L.control.zoom({{position:'topright'}}).addTo(map);
L.control.attribution({{position:'bottomright',prefix:false}})
  .addAttribution('© <a href="https://carto.com">CARTO</a>')
  .addTo(map);

// ── Sidebar ────────────────────────────────────────────────────────────────────
let activeCid = null;
const sidebar = $('sidebar');

function openSidebar(cid){{
  activeCid = cid;
  renderSidebar(cid);
  sidebar.classList.add('open');
  // Nudge map west so pin isn't hidden behind sidebar
  map.panBy([180, 0], {{animate:true, duration:.25}});
}}

function closeSidebar(){{
  sidebar.classList.remove('open');
  if(activeCid){{
    map.panBy([-180, 0], {{animate:true, duration:.25}});
    activeCid = null;
  }}
}}

$('sb-close').addEventListener('click', closeSidebar);
map.on('click', () => {{ if(activeCid) closeSidebar(); }});

// ── Render sidebar content ─────────────────────────────────────────────────────
function scoreDotsHtml(score){{
  let h='<div class="pref-score">';
  for(let i=1;i<=3;i++)
    h+=`<div class="pref-dot ${{i<=score?'dot-on':'dot-off'}}"></div>`;
  return h+'</div>';
}}

function renderSidebar(cid){{
  const c = COLLECTORS[cid];
  const todTotal = c.tod.AM + c.tod.MD + c.tod.PM;
  const DAYS = ['Mon','Tue','Wed','Thu','Fri','Sat','Sun'];

  // ── Stats ──────────────────────────────────────────────────────────────────
  const statsHtml = `
    <div class="stats-row">
      <div class="stat-card">
        <div class="stat-val">${{c.total}}</div>
        <div class="stat-lbl">Total</div>
      </div>
      <div class="stat-card">
        <div class="stat-val">${{c.month}}</div>
        <div class="stat-lbl">This Month</div>
      </div>
      <div class="stat-card">
        <div class="stat-val">${{c.week}}</div>
        <div class="stat-lbl">This Week</div>
      </div>
    </div>`;

  // ── TOD breakdown ──────────────────────────────────────────────────────────
  let todHtml = '';
  if(todTotal > 0){{
    todHtml = `<div class="section">
      <div class="sec-title">Time of Day</div>
      ${{['AM','MD','PM'].map(tod=>{{
        const pct = todTotal > 0 ? Math.round(c.tod[tod]/todTotal*100) : 0;
        return `<div class="tod-bar-row">
          <span class="tod-lbl">${{tod}}</span>
          <div class="tod-bg"><div class="tod-fill" style="width:${{pct}}%;background:${{TOD_COLORS[tod]}}"></div></div>
          <span class="tod-num">${{c.tod[tod]}}</span>
        </div>`;
      }}).join('')}}
    </div>`;
  }}

  // ── Availability ───────────────────────────────────────────────────────────
  let availHtml = '';
  if(Object.keys(c.avail).length > 0){{
    availHtml = `<div class="section">
      <div class="sec-title">Availability</div>
      <div class="avail-grid">
        <div class="ag-head"></div>
        ${{DAYS.map(d=>`<div class="ag-head">${{d.slice(0,2)}}</div>`).join('')}}
        ${{['AM','MD','PM'].map(tod=>{{
          const row = c.avail[tod] || Array(7).fill(false);
          return `<div class="ag-tod">${{tod}}</div>
            ${{row.map(v=>`<div class="ag-cell ${{v?'ag-yes':'ag-no'}}">${{v?'✓':''}}</div>`).join('')}}`;
        }}).join('')}}
      </div>
    </div>`;
  }}

  // ── Route preferences ──────────────────────────────────────────────────────
  let prefHtml = '';
  if(c.pref && c.pref.length > 0){{
    prefHtml = `<div class="section">
      <div class="sec-title">Route Preferences</div>
      ${{c.pref.map(p=>`
        <div class="pref-row">
          <span class="pref-name">${{p.name}}</span>
          ${{scoreDotsHtml(p.score)}}
        </div>`).join('')}}
    </div>`;
  }}

  // ── Recent walks ───────────────────────────────────────────────────────────
  let recentHtml = '';
  if(c.recent && c.recent.length > 0){{
    recentHtml = `<div class="section">
      <div class="sec-title">Recent Walks</div>
      ${{c.recent.map(w=>`
        <div class="walk-row">
          <span class="tod-badge tod-${{w.tod}}">${{w.tod}}</span>
          <span class="walk-route">${{ROUTES_META[w.route]?.name || w.route}}</span>
          <span class="walk-date">${{fmtDate(w.date)}}</span>
        </div>`).join('')}}
    </div>`;
  }} else {{
    recentHtml = `<div class="section">
      <div class="sec-title">Recent Walks</div>
      <div style="font-size:12px;color:#6b7099;padding:4px 0">No completed walks yet</div>
    </div>`;
  }}

  // ── Upcoming scheduled ─────────────────────────────────────────────────────
  let upcomingHtml = '';
  if(c.upcoming && c.upcoming.length > 0){{
    upcomingHtml = `<div class="section">
      <div class="sec-title">Upcoming Scheduled</div>
      ${{c.upcoming.map(a=>`
        <div class="walk-row">
          <span class="tod-badge tod-${{a.tod}}">${{a.tod}}</span>
          <span class="walk-route">${{ROUTES_META[a.route]?.name || a.route}}</span>
          <span class="walk-date">${{fmtDate(a.date)}}</span>
          ${{a.backpack?`<span class="walk-bp">BP ${{a.backpack}}</span>`:''}}
        </div>`).join('')}}
    </div>`;
  }}

  $('sb-content').innerHTML = `
    <div class="sb-head" style="--c:${{c.color}}">
      <div class="sb-avatar" style="--c:${{c.color}}">${{cid}}</div>
      <div>
        <div class="sb-name">${{c.name}}</div>
        <div class="sb-id">${{cid}}</div>
      </div>
    </div>
    ${{statsHtml}}
    ${{todHtml}}
    ${{availHtml}}
    ${{prefHtml}}
    ${{recentHtml}}
    ${{upcomingHtml}}
  `;
}}

// ── Create SVG pin icon ────────────────────────────────────────────────────────
function makePinIcon(color, count){{
  const sz  = count > 9 ? 42 : 36;
  const fsz = count > 9 ? 12 : 13;
  const svg = `<svg xmlns="http://www.w3.org/2000/svg" width="${{sz}}" height="${{Math.round(sz*1.25)}}"
      viewBox="0 0 36 45">
    <path d="M18 1C8.6 1 1 8.6 1 18C1 29.5 18 44 18 44S35 29.5 35 18C35 8.6 27.4 1 18 1Z"
      fill="${{color}}" stroke="rgba(255,255,255,.25)" stroke-width="1.2"/>
    <circle cx="18" cy="18" r="11.5" fill="rgba(0,0,0,.22)"/>
    <text x="18" y="23" text-anchor="middle" fill="white"
      font-family="-apple-system,BlinkMacSystemFont,sans-serif"
      font-weight="800" font-size="${{fsz}}">${{count}}</text>
  </svg>`;
  return L.divIcon({{
    html: svg,
    className: 'cpin-icon',
    iconSize:   [sz, Math.round(sz*1.25)],
    iconAnchor: [sz/2, Math.round(sz*1.25)],
  }});
}}

// ── Place markers ──────────────────────────────────────────────────────────────
const markers = {{}};

Object.values(COLLECTORS).forEach(c => {{
  const icon   = makePinIcon(c.color, c.total);
  const marker = L.marker([c.lat, c.lng], {{icon, title:c.name}}).addTo(map);

  marker.on('click', e => {{
    L.DomEvent.stopPropagation(e);
    if(activeCid === c.id) {{ closeSidebar(); return; }}
    openSidebar(c.id);
  }});

  // Hover tooltip
  marker.bindTooltip(`<b>${{c.name}}</b><br>${{c.total}} walk${{c.total!==1?'s':''}} completed`, {{
    direction:'top', offset:[0,-38], className:'leaflet-tooltip',
    opacity:.95,
  }});

  markers[c.id] = marker;
}});

// ── Legend ─────────────────────────────────────────────────────────────────────
const legBody = $('leg-body');
Object.values(COLLECTORS).forEach(c => {{
  const row = document.createElement('div');
  row.className = 'leg-row';
  row.innerHTML = `
    <div class="leg-dot" style="background:${{c.color}}"></div>
    <span class="leg-name">${{c.name}}</span>
    <span class="leg-count">${{c.total}}</span>`;
  row.addEventListener('click', () => {{
    map.setView([c.lat, c.lng], 12, {{animate:true}});
    openSidebar(c.id);
  }});
  legBody.appendChild(row);
}});
</script>
</body>
</html>"""

OUT.parent.mkdir(parents=True, exist_ok=True)
OUT.write_text(HTML, encoding="utf-8")
print(f"OK Written: {OUT}")
print(f"   Collectors: {len(collector_payload)}")
print(f"   Walks logged: {len(walks)}")
print(f"   Scheduled assignments: {len(sched_assignments)}")
