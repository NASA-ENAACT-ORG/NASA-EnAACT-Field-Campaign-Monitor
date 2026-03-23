#!/usr/bin/env python3
"""
walk_scheduler.py — NYC Field Data Campaign Walk Scheduler
============================================================
Generates a ranked list of top-8 recommended walks and a weekly calendar
assignment for the current forecast week.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
SETUP (one-time)
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

1. Install dependencies (run once in any terminal):

       pip install anthropic pdfplumber openpyxl pandas

2. Get your Anthropic API key from https://console.anthropic.com
   and set it as an environment variable for this session:

   Windows Command Prompt:
       set ANTHROPIC_API_KEY=sk-ant-xxxxx

   Windows PowerShell:
       $env:ANTHROPIC_API_KEY="sk-ant-xxxxx"

   Mac / Linux / Git Bash:
       export ANTHROPIC_API_KEY=sk-ant-xxxxx

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
HOW TO RUN
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

Open a terminal, cd into this folder, then run:

   python walk_scheduler.py

The script reads all files from the current directory
(Walks_Log.txt, Preferred_Routes.xlsx, Forecast/,
Collector_Schedule/, Route_KMLs/) automatically — no arguments needed.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
UPDATING WEEKLY DATA
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

• New forecast week: drop the PDF directly into Forecast/ with a
  filename like "Mar 15 - Mar 21.pdf" (revision suffix optional, e.g.
  "Mar 15 - Mar 21 R2.pdf").  The script picks the most recently
  modified file whose week end-date has not yet fully passed —
  no subfolders or revision numbering required.

• New walk completions: append lines to Walks_Log.txt using either
  the old format  Backpack_Collector_Boro_Neigh_MM_DD_YYYY_TOD
  e.g.  A_SOT_MN_MT_03_14_2026_AM
  or the new format  Backpack_Collector_Boro_Neigh_YYYYMMDD_TOD
  e.g.  X_SOT_MN_HT_20260309_MD   (X = legacy pre-A/B backpack code)

• Recalibration completed: append a line to Walks_Log.txt in
  the format  RECAL_MM_DD_YYYY  (e.g. RECAL_03_14_2026).
  The scheduler proposes a recal day once 25 days have elapsed since
  the last logged recal (targeting ~30-day spacing).  It picks the
  worst-weather business day in the current forecast window; if the
  entire week is good weather it falls back to the least-productive
  field day to avoid missing the 30-day target.

• New collector schedules: drop a screenshot or PDF into
  Collector_Schedule/ with the collector's name in the filename
  (e.g. "Screenshot - Tahani.png"). The script will detect and
  parse it automatically via Claude vision.
"""

import argparse
import os
import re
import sys
import json
import math

# Force UTF-8 output on Windows so box-drawing characters render correctly
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
import base64
import shutil
from datetime import date, datetime, timedelta
from pathlib import Path
from collections import defaultdict
from typing import Dict, List, Optional, Tuple

import anthropic
import pdfplumber
import pandas as pd
import xml.etree.ElementTree as ET
import folium

# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────────────────────────────────────────

BASE_DIR      = Path(__file__).parent
WALKS_LOG     = BASE_DIR / "Walks_Log.txt"
PREF_ROUTES    = BASE_DIR / "Preferred_Routes.xlsx"
PREF_ROUTES_V2 = BASE_DIR / "V2_Preferred_Routes.xlsx"
FORECAST_DIR  = BASE_DIR / "Forecast"
SCHEDULE_DIR  = BASE_DIR / "Collector_Schedule"
KML_DIR       = BASE_DIR / "Route_KMLs"

CURRENT_SEASON      = "Spring"
CLOUD_THRESHOLD     = 33          # ≤ this % cloud cover = good weather
TARGET_COMPLETIONS  = 8
MIN_COMPLETIONS     = 6
TODS                = ["AM", "MD", "PM"]
CURRENT_YEAR        = date.today().year    # used for date inference throughout
CLAUDE_MODEL        = "claude-haiku-4-5-20251001"

# CCNY (Marshak Science Building, 160 Convent Ave) — home base for recalibration
CCNY_LAT =  40.8196
CCNY_LON = -73.9496

# LaGuardia Community College (31-10 Thomson Ave, Long Island City)
LAGCC_LAT =  40.7425
LAGCC_LON = -73.9236

# Campus-to-backpack binding
# Backpack A is tied to CCNY; Backpack B is tied to LAGCC.
# Routes are assigned to whichever campus is closer by transit time.
CAMPUS_COORDS = {
    "A": (CCNY_LAT, CCNY_LON, "CCNY"),
    "B": (LAGCC_LAT, LAGCC_LON, "LAGCC"),
}
# Proxy routes for campus transit-time lookups
# (CCNY ≈ 135 St / MN_HT stop; LAGCC ≈ Court Sq / QN_LA stop)
CAMPUS_PROXY_ROUTE = {"A": "MN_HT", "B": "QN_LA"}

# Student collectors per backpack (drive scheduling)
BACKPACK_COLLECTORS: Dict[str, set] = {
    "A": {"JEN", "AYA", "SOT", "TAH"},   # CCNY team
    "B": {"TER", "ALX", "SCT", "JAM"},   # LaGCC team
}
# ANG is CCNY-affiliated staff; only used as a last resort for Backpack A
LAST_RESORT_COLLECTORS = ["ANG"]
LAST_RESORT_BACKPACK   = "A"

# Professors / support staff — excluded from regular scheduling
STAFF_COLLECTORS = ["NRS", "PRA", "NAT"]

# Full collector list (students + ANG; excludes staff)
COLLECTORS = ["SOT", "AYA", "ALX", "TAH", "JAM", "JEN", "SCT", "TER", "ANG"]

# Map collector IDs to the first-name used in Collector_Locs.kml
COLLECTOR_KML_NAMES = {
    "SOT": "Soteri",
    "AYA": "Aya",
    "ALX": "Alex",
    "JAM": "James",
    "JEN": "Jennifer",
    "SCT": "Scott",
    "TER": "Terra",
    "ANG": "Angy",
    "TAH": "Taha",
    "NRS": "Prof. Naresh Devineni",
    "PRA": "Prof. Prathap Ramamurthy",
}
COLLECTOR_ID_TO_NAME = COLLECTOR_KML_NAMES

# Route definitions: boro code → neighbourhood codes
ROUTES: Dict[str, List[str]] = {
    "MN": ["HT", "WH", "UE", "MT", "LE"],
    "BX": ["HP", "NW"],
    "BK": ["DT", "WB", "BS", "CH", "SP", "CI"],
    "QN": ["FU", "LI", "JH", "JA", "FH", "LA", "EE"],
}
ALL_ROUTES = [f"{b}_{n}" for b, ns in ROUTES.items() for n in ns]

# Human-readable route labels
ROUTE_LABELS = {
    "MN_HT": "Manhattan – Harlem",
    "MN_WH": "Manhattan – Washington Hts",
    "MN_UE": "Manhattan – Upper East Side",
    "MN_MT": "Manhattan – Midtown",
    "MN_LE": "Manhattan – Union Sq/LES",
    "BX_HP": "Bronx – Hunts Point",
    "BX_NW": "Bronx – Norwood",
    "BK_DT": "Brooklyn – Downtown BK",
    "BK_WB": "Brooklyn – Williamsburg",
    "BK_BS": "Brooklyn – Bed Stuy",
    "BK_CH": "Brooklyn – Crown Heights",
    "BK_SP": "Brooklyn – Sunset Park",
    "BK_CI": "Brooklyn – Coney Island",
    "QN_FU": "Queens – Flushing",
    "QN_LI": "Queens – Astoria/LIC",
    "QN_JH": "Queens – Jackson Heights",
    "QN_JA": "Queens – Jamaica",
    "QN_FH": "Queens – Forest Hills",
    "QN_LA": "Queens – LaGuardia CC",
    "QN_EE": "Queens – East Elmhurst",
}

# KML name used in each boro file → route code
KML_NAME_TO_ROUTE = {
    "Harlem":                      "MN_HT",
    "Washington Heights":          "MN_WH",
    "Upper East Side":             "MN_UE",
    "Midtown":                     "MN_MT",
    "Union Square/LES":            "MN_LE",
    "Norwood":                     "BX_NW",
    "Hunts Point":                 "BX_HP",
    "Downtown Brooklyn":           "BK_DT",
    "Williamsburg":                "BK_WB",
    "Bed Sty":                     "BK_BS",
    "Crown Heights":               "BK_CH",
    "Sunset Park":                 "BK_SP",
    "Coney Island":                "BK_CI",
    "Flushing":                    "QN_FU",
    "Astoria/LIC":                 "QN_LI",
    "Jackson Heights":             "QN_JH",
    "Jamaica":                     "QN_JA",
    "Forest Hills":                "QN_FH",
    "LaGuardia Community College": "QN_LA",
    "East Elmhurst":               "QN_EE",
}

# Substring → collector ID for schedule filename matching
FILENAME_TO_COLLECTOR = {
    "terra": "TER", "emmerich": "TER",
    "aya":   "AYA", "nasri":    "AYA",
    "scott": "SCT", "atlixqueno": "SCT",
    "alex":  "ALX", "leon":     "ALX",
    "james": "JAM", "lu":       "JAM",
    "jennifer": "JEN", "ramirez": "JEN",
    "soteri": "SOT",
    "pra":   "PRA", "prathap": "PRA",
    "nat":   "NAT", "natalie": "NAT",
    "nrs":   "NRS",
    "tah":   "TAH", "tahani":  "TAH",
}

# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def get_season(d: date) -> str:
    """Meteorological season for a given date."""
    m = d.month
    if   m in (3, 4, 5):  return "Spring"
    elif m in (6, 7, 8):  return "Summer"
    elif m in (9, 10, 11): return "Fall"
    else:                  return "Winter"


def haversine_km(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    """Great-circle distance between two lat/lon points in kilometres."""
    R = 6371.0
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dp = math.radians(lat2 - lat1)
    dl = math.radians(lon2 - lon1)
    a = math.sin(dp / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dl / 2) ** 2
    return R * 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))


def image_to_base64(filepath: Path) -> Tuple[str, str]:
    """Return (base64_string, media_type) for an image file."""
    ext = filepath.suffix.lower()
    media_types = {
        ".png": "image/png", ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg", ".gif": "image/gif", ".webp": "image/webp",
    }
    media_type = media_types.get(ext, "image/png")
    with open(filepath, "rb") as f:
        data = base64.standard_b64encode(f.read()).decode("utf-8")
    return data, media_type


def parse_week_folder_name(name: str) -> Tuple[Optional[date], Optional[date]]:
    """
    Parse folder names like 'Mar 8 - Mar 14', 'Feb 22 - Mar 2' → (start, end).
    Year is inferred as 2026.
    """
    MONTHS = {
        "Jan":1,"Feb":2,"Mar":3,"Apr":4,"May":5,"Jun":6,
        "Jul":7,"Aug":8,"Sep":9,"Oct":10,"Nov":11,"Dec":12
    }
    # Split on ' - '
    halves = re.split(r"\s+-\s+", name.strip(), maxsplit=1)
    if len(halves) < 2:
        return None, None

    def _parse_part(s: str) -> Optional[date]:
        m = re.match(r"([A-Za-z]+)\s+(\d+)", s.strip())
        if m and m.group(1) in MONTHS:
            return date(CURRENT_YEAR, MONTHS[m.group(1)], int(m.group(2)))
        return None

    return _parse_part(halves[0]), _parse_part(halves[1])


def safe_copy(src: Path, suffix: str) -> Path:
    """Copy a file to the system temp dir to avoid OneDrive locking."""
    tmp = Path(os.environ.get("TEMP", "/tmp")) / f"walk_sched_{suffix}"
    shutil.copy2(src, tmp)
    return tmp


# ─────────────────────────────────────────────────────────────────────────────
# STEP 1 — Parse Walks Log
# ─────────────────────────────────────────────────────────────────────────────

def _parse_walk_date(parts: List[str], line: str) -> Optional[date]:
    """
    Extract a date from a split walk-code.

    Supported formats
    -----------------
    Old (8 parts):  BP_COL_BORO_NEIGH_MM_DD_YYYY_TOD  → parts[4..6]
    New (6 parts):  BP_COL_BORO_NEIGH_YYYYMMDD_TOD    → parts[4]
    """
    try:
        if len(parts) == 8:
            mm, dd, yyyy = parts[4], parts[5], parts[6]
            return date(int(yyyy), int(mm), int(dd))
        elif len(parts) == 6:
            raw = parts[4]          # e.g. "20260309"
            return date(int(raw[:4]), int(raw[4:6]), int(raw[6:8]))
    except (ValueError, IndexError):
        pass
    print(f"  [WARN] Invalid date in walk code: {line}")
    return None


def parse_walks_log() -> Dict[Tuple[str, str, str], int]:
    """
    Parse Walks_Log.txt and count completions per (route, tod, season).

    Accepted walk-code formats
    --------------------------
    Old (8 parts):  BP_COL_BORO_NEIGH_MM_DD_YYYY_TOD   BP ∈ {A, B, X}
    New (6 parts):  BP_COL_BORO_NEIGH_YYYYMMDD_TOD      BP ∈ {A, B, X}

    Backpack X is a legacy code pre-dating A/B; walks count toward
    completion totals identically to A or B.
    """
    VALID_BACKPACKS = {"A", "B", "X"}
    completions: Dict[Tuple[str, str, str], int] = defaultdict(int)
    seen_entries: set = set()

    try:
      f_open = open(WALKS_LOG, encoding="utf-8")
    except FileNotFoundError:
        print(f"  [WARN] {WALKS_LOG.name} not found — no walk history loaded")
        return {}

    with f_open as f:
        for raw in f:
            line = raw.strip()
            if not line or line.startswith("#"):
                continue

            parts = line.split("_")
            if len(parts) not in (6, 8):
                print(f"  [WARN] Malformed walk code (skipping): {line}")
                continue

            bp = parts[0]
            if bp not in VALID_BACKPACKS:
                print(f"  [WARN] Unknown backpack '{bp}' in: {line}")
                continue

            boro  = parts[2]
            neigh = parts[3]
            tod   = parts[-1]       # always the last field in both formats

            walk_date = _parse_walk_date(parts, line)
            if walk_date is None:
                continue

            route  = f"{boro}_{neigh}"
            season = get_season(walk_date)

            if route not in ALL_ROUTES:
                print(f"  [WARN] Unknown route '{route}' in: {line}")
                continue
            if tod not in TODS:
                print(f"  [WARN] Unknown TOD '{tod}' in: {line}")
                continue

            entry_key = (parts[0], parts[1], route, walk_date, tod)
            if entry_key in seen_entries:
                print(f"  [WARN] Duplicate walk entry skipped: {line}")
                continue
            seen_entries.add(entry_key)
            completions[(route, tod, season)] += 1

    return dict(completions)


# ─────────────────────────────────────────────────────────────────────────────
# STEP 2 — Parse Current-Week Forecast PDF
# ─────────────────────────────────────────────────────────────────────────────

def find_current_week_forecast() -> Tuple[Path, date, date]:
    """
    Scan Forecast/ directly (flat — no subfolders) for all PDFs.
    Keep only files whose parsed week end-date >= today so that weeks
    that have fully passed are ignored.
    Among the remaining candidates pick the most recently modified file;
    a freshly saved PDF always wins over an older one regardless of name.
    """
    import datetime as _dt
    today = date.today()
    candidates = []

    for p in FORECAST_DIR.glob("*.pdf"):
        start, end = parse_week_folder_name(p.stem)
        if start is None or end is None:
            continue
        if end < today:
            continue                        # week fully in the past
        candidates.append((p.stat().st_mtime, start, end, p))

    if not candidates:
        raise FileNotFoundError(
            "No forecast PDFs found in Forecast/ whose week end-date >= today. "
            "Drop a PDF named like 'Mar 15 - Mar 21.pdf' into the Forecast/ folder."
        )

    candidates.sort(key=lambda x: x[0], reverse=True)   # newest first
    mtime, start, end, best_pdf = candidates[0]
    mtime_str = _dt.datetime.fromtimestamp(mtime).strftime("%Y-%m-%d %H:%M")
    print(f"  Using forecast: {best_pdf.name}  (last modified {mtime_str})")
    return best_pdf, start, end


def parse_forecast_pdf(pdf_path: Path) -> Dict[Tuple[date, str, str], bool]:
    """
    Extract cloud-coverage data from the forecast PDF.

    Supports two PDF formats:
      • City-wide (13 cols): [DayName\\nDate, 7am, 8am, 9am, AM_avg, 12pm, 1pm, 2pm,
                               MD_avg, 5pm, 6pm, 7pm, PM_avg]
        → stored with boro key ""
      • Per-borough (14 cols): [Borough, DayName\\nDate, 7am, 8am, 9am, AM_avg, 12pm,
                                 1pm, 2pm, MD_avg, 5pm, 6pm, 7pm, PM_avg]
        → stored with boro key BX / MN / QN / BK; borough cell may be None (merged)

    Returns {(date, tod, boro): True if avg ≤ CLOUD_THRESHOLD}
    Boro is "" when the PDF has no per-borough breakdown.
    """
    BORO_KEYWORDS = {
        "bronx": "BX", "manhattan": "MN", "queens": "QN", "brooklyn": "BK",
    }

    tmp = safe_copy(pdf_path, "forecast.pdf")
    weather: Dict[Tuple[date, str, str], bool] = {}
    current_boro = ""   # updated as we scan rows

    def pct(cell) -> Optional[int]:
        if cell is None:
            return None
        m = re.search(r"(\d+)", str(cell))
        return int(m.group(1)) if m else None

    with pdfplumber.open(tmp) as pdf:
        for page in pdf.pages:
            for table in page.extract_tables():
                for row in table:
                    if not row:
                        continue

                    # ── Detect borough header in any non-None cell ──────────
                    for cell in row:
                        if cell is None:
                            continue
                        lower = str(cell).strip().lower()
                        for kw, code in BORO_KEYWORDS.items():
                            if kw in lower:
                                current_boro = code
                                break

                    # ── Find which cell holds the date ──────────────────────
                    date_col = None
                    d = None
                    for ci, cell in enumerate(row):
                        if cell is None:
                            continue
                        dm = re.search(r"(\d+)/(\d+)/(\d{2,4})", str(cell))
                        if dm:
                            try:
                                mo_ = int(dm.group(1))
                                dy_ = int(dm.group(2))
                                yr_ = int(dm.group(3))
                                yr_ = (2000 + yr_) if yr_ < 100 else yr_
                                d = date(yr_, mo_, dy_)
                                date_col = ci
                            except ValueError:
                                pass
                            break
                    if d is None or date_col is None:
                        continue

                    # ── Column offsets depend on format width ───────────────
                    # 13-col: date at 0 → avgs at 4, 8, 12
                    # 14-col: date at 1 → avgs at 5, 9, 13
                    offset = date_col  # 0 → city-wide, 1 → per-borough
                    need = 13 + offset
                    if len(row) < need:
                        continue

                    am_avg = pct(row[4  + offset])
                    md_avg = pct(row[8  + offset])
                    pm_avg = pct(row[12 + offset])

                    boro_key = current_boro if offset > 0 else ""
                    if am_avg is not None:
                        weather[(d, "AM", boro_key)] = am_avg <= CLOUD_THRESHOLD
                    if md_avg is not None:
                        weather[(d, "MD", boro_key)] = md_avg <= CLOUD_THRESHOLD
                    if pm_avg is not None:
                        weather[(d, "PM", boro_key)] = pm_avg <= CLOUD_THRESHOLD

    if not weather:
        print("  [WARN] Table parsing failed; falling back to raw-text extraction.")
        weather = _parse_forecast_text_fallback(tmp)

    # Log detected boroughs
    boros_found = sorted({b for (_, _, b) in weather if b})
    if boros_found:
        print(f"  [forecast] Per-borough weather loaded: {', '.join(boros_found)}")
    else:
        print("  [forecast] City-wide (no per-borough breakdown) weather loaded.")

    return weather


def _parse_forecast_text_fallback(pdf_path: Path) -> Dict[Tuple[date, str, str], bool]:
    """
    Regex-based fallback: match lines like
    'Sunday\\n3/8/26  98% 99% 99% 99%  85% 80% 86% 84%  40% 16% 11% 22%'
    Returns city-wide (boro="") entries.
    """
    weather: Dict[Tuple[date, str, str], bool] = {}
    with pdfplumber.open(pdf_path) as pdf:
        text = "\n".join(p.extract_text() or "" for p in pdf.pages)

    # Build a single flat string to match across newlines
    flat = re.sub(r"\s+", " ", text)
    pattern = re.compile(
        r"(?:Sunday|Monday|Tuesday|Wednesday|Thursday|Friday|Saturday)\s+"
        r"(\d+)/(\d+)/(\d+)\s+"
        r"(\d+)%\s+\d+%\s+\d+%\s+(\d+)%\s+"   # 7am 8am 9am AM_avg
        r"\d+%\s+\d+%\s+\d+%\s+(\d+)%\s+"      # 12pm 1pm 2pm MD_avg
        r"\d+%\s+\d+%\s+\d+%\s+(\d+)%"         # 5pm 6pm 7pm PM_avg
    )
    for m in pattern.finditer(flat):
        mo, dy, yr_ = int(m.group(1)), int(m.group(2)), int(m.group(3))
        yr = (2000 + yr_) if yr_ < 100 else yr_
        try:
            d = date(yr, mo, dy)
        except ValueError:
            continue
        weather[(d, "AM", "")] = int(m.group(5)) <= CLOUD_THRESHOLD
        weather[(d, "MD", "")] = int(m.group(6)) <= CLOUD_THRESHOLD
        weather[(d, "PM", "")] = int(m.group(7)) <= CLOUD_THRESHOLD
    return weather


# ─────────────────────────────────────────────────────────────────────────────
# STEP 3 — Parse Preferred Routes
# ─────────────────────────────────────────────────────────────────────────────

def parse_preferred_routes() -> Dict[str, Dict[str, int]]:
    """
    Parse route affinity scores per collector.
    Uses V2_Preferred_Routes.xlsx (0–3 scale) where data exists for a collector,
    falling back to Preferred_Routes.xlsx (binary 1 → score 1) for the rest.

    Returns {collector_id: {neigh_code: score}}
      score 0 = explicitly do not assign
      score 1 = least desirable but willing
      score 2 = acceptable
      score 3 = most desirable
    Collectors with no entry for a route are treated as neutral (ok to assign).
    """
    affinity: Dict[str, Dict[str, int]] = {c: {} for c in COLLECTORS}
    all_neigh_codes: set = {n for ns in ROUTES.values() for n in ns}
    v2_collectors: set  = set()

    # ── V2: rows = collectors, columns = route codes, values = 0–3 ──────────
    if PREF_ROUTES_V2.exists():
        tmp = safe_copy(PREF_ROUTES_V2, "v2_preferred_routes.xlsx")
        df  = pd.read_excel(tmp, header=None)

        # Find header row: contains the most known route codes
        route_row_idx = -1
        col_to_route: Dict[int, str] = {}
        for i, row in df.iterrows():
            found = {
                j: str(v).strip()
                for j, v in enumerate(row)
                if str(v).strip() in all_neigh_codes
            }
            if len(found) > len(col_to_route):
                route_row_idx = i
                col_to_route  = found

        if route_row_idx < 0:
            print("  [WARN] Could not locate route header row in V2_Preferred_Routes.xlsx")
        else:
            for i in range(route_row_idx + 1, len(df)):
                row = df.iloc[i]
                cid: Optional[str] = None
                for v in row:
                    if str(v).strip() in COLLECTORS:
                        cid = str(v).strip()
                        break
                if cid is None:
                    continue
                v2_collectors.add(cid)
                for col_j, neigh_code in col_to_route.items():
                    if col_j < len(row):
                        val = row.iloc[col_j]
                        if pd.notna(val):
                            try:
                                score = int(float(str(val).strip()))
                                affinity[cid][neigh_code] = score  # store 0 = explicit exclusion
                            except (ValueError, TypeError):
                                pass
    else:
        print("  [WARN] V2_Preferred_Routes.xlsx not found — using V1 only")

    # ── V1 fallback: binary 1 → score 1, only for collectors absent from V2 ─
    if PREF_ROUTES.exists():
        tmp = safe_copy(PREF_ROUTES, "preferred_routes.xlsx")
        df  = pd.read_excel(tmp, header=None)

        collector_row_idx = -1
        col_to_collector: Dict[int, str] = {}
        for i, row in df.iterrows():
            found_cols = {
                j: str(v).strip()
                for j, v in enumerate(row)
                if str(v).strip() in COLLECTORS
            }
            if len(found_cols) > len(col_to_collector):
                collector_row_idx = i
                col_to_collector  = found_cols

        if collector_row_idx < 0:
            print("  [WARN] Could not locate collector header row in Preferred_Routes.xlsx")
        else:
            for i in range(collector_row_idx + 1, len(df)):
                row = df.iloc[i]
                neigh_code: Optional[str] = None
                for v in row:
                    if str(v).strip() in all_neigh_codes:
                        neigh_code = str(v).strip()
                        break
                if neigh_code is None:
                    continue
                for col_j, cid in col_to_collector.items():
                    if cid in v2_collectors:
                        continue  # V2 takes precedence for this collector
                    if col_j < len(row):
                        val = row.iloc[col_j]
                        if pd.notna(val) and str(val).strip() == "1":
                            affinity[cid][neigh_code] = 1  # binary → score 1

    v1_count = sum(1 for c in affinity if affinity[c] and c not in v2_collectors)
    print(f"  V2 scores: {len(v2_collectors)} collectors | V1 fallback: {v1_count} collectors")
    return affinity


# ─────────────────────────────────────────────────────────────────────────────
# STEP 4 — Parse Collector Schedules (Claude Vision)
# ─────────────────────────────────────────────────────────────────────────────

_SCHEDULE_PROMPT = """\
You are extracting walk availability from a collector's calendar screenshot or schedule document.

Walk time-of-day windows (all times local):
  • AM  = 7:00 AM – 10:00 AM
  • MD  = 11:30 AM –  2:30 PM
  • PM  =  4:30 PM –  7:30 PM

A collector is AVAILABLE for a TOD only if they have NO commitment overlapping that window.
Be conservative: if a class or job partially overlaps a window, mark it unavailable.

The schedule may show a specific week OR be a recurring semester schedule.
If it shows specific calendar dates (e.g. "Mar 9", "Mon Mar 9"), use schedule_type "specific"
and record each date's availability. Otherwise use schedule_type "recurring" and record by
day-of-week key (MON, TUE, WED, THU, FRI, SAT, SUN).

If the document contains an explicit availability note (e.g. "Mornings: Mon/Tue"), trust it.

Return ONLY valid JSON — no markdown, no extra text — in this exact structure:
{
  "schedule_type": "specific",
  "week_start_date": "2026-03-09",
  "availability": {
    "2026-03-09": {"AM": true,  "MD": true,  "PM": false},
    "2026-03-10": {"AM": true,  "MD": false, "PM": false}
  },
  "notes": "..."
}
OR
{
  "schedule_type": "recurring",
  "week_start_date": null,
  "availability": {
    "MON": {"AM": true,  "MD": false, "PM": true},
    "TUE": {"AM": true,  "MD": false, "PM": true},
    "WED": {"AM": false, "MD": false, "PM": false},
    "THU": {"AM": true,  "MD": true,  "PM": false},
    "FRI": {"AM": true,  "MD": false, "PM": true},
    "SAT": {"AM": true,  "MD": true,  "PM": true},
    "SUN": {"AM": true,  "MD": true,  "PM": true}
  },
  "notes": "..."
}
"""


def _extract_json(text: str) -> Optional[Dict]:
    """Pull the first {...} block from a string and parse as JSON."""
    m = re.search(r"\{.*\}", text, re.DOTALL)
    if not m:
        return None
    try:
        return json.loads(m.group())
    except json.JSONDecodeError as e:
        print(f"    [WARN] JSON parse error: {e}")
        return None


def _call_vision(client: anthropic.Anthropic, image_data: str, media_type: str) -> Optional[Dict]:
    try:
        resp = client.messages.create(
            model=CLAUDE_MODEL,
            max_tokens=1024,
            messages=[{
                "role": "user",
                "content": [
                    {"type": "image",
                     "source": {"type": "base64", "media_type": media_type, "data": image_data}},
                    {"type": "text", "text": _SCHEDULE_PROMPT},
                ],
            }],
        )
        return _extract_json(resp.content[0].text)
    except anthropic.APIError as e:
        print(f"    [WARN] Claude API error in vision call: {e}")
        return None


def _call_text(client: anthropic.Anthropic, text: str) -> Optional[Dict]:
    try:
        resp = client.messages.create(
            model=CLAUDE_MODEL,
            max_tokens=1024,
            messages=[{
                "role": "user",
                "content": f"{_SCHEDULE_PROMPT}\n\nSchedule text to parse:\n{text}",
            }],
        )
        return _extract_json(resp.content[0].text)
    except anthropic.APIError as e:
        print(f"    [WARN] Claude API error in text call: {e}")
        return None


def parse_collector_schedules(
    client: anthropic.Anthropic,
) -> Dict[str, Dict]:
    """
    For each file in Collector_Schedule/, identify the collector, send the
    image/PDF to Claude for vision-based parsing, and collect results.
    Returns {collector_id: parsed_schedule_dict}
    """
    schedules: Dict[str, Dict] = {}

    for filepath in sorted(SCHEDULE_DIR.iterdir()):
        if filepath.name.startswith("."):
            continue
        # Identify collector from filename
        fname_lower = filepath.name.lower()
        cid: Optional[str] = None
        for key, c in FILENAME_TO_COLLECTOR.items():
            if key in fname_lower:
                cid = c
                break
        if cid is None:
            print(f"  [WARN] Cannot map file to a collector: {filepath.name}")
            continue

        print(f"  Parsing {cid} schedule: {filepath.name}")
        result: Optional[Dict] = None

        try:
            if filepath.suffix.lower() == ".pdf":
                # Try extracting text first (cheaper); fall back to vision not possible for PDFs
                tmp = safe_copy(filepath, f"sched_{cid}.pdf")
                with pdfplumber.open(tmp) as pdf:
                    text = "\n".join(p.extract_text() or "" for p in pdf.pages)
                if text.strip():
                    result = _call_text(client, text)
                # If PDF has no extractable text, we skip (PDF-to-image conversion
                # would require extra libs; the text extraction above covers our case)
            else:
                # Image file
                img_data, media_type = image_to_base64(filepath)
                result = _call_vision(client, img_data, media_type)
        except Exception as e:
            print(f"    [ERROR] {e}")

        if result:
            schedules[cid] = result
            avail = result.get("availability", {})
            print(f"    → schedule_type={result.get('schedule_type')} "
                  f"keys={list(avail.keys())[:5]}")
        else:
            print(f"    [WARN] No usable schedule extracted for {cid}")

    # Collectors without any schedule file → assume fully available
    for cid in COLLECTORS:
        if cid not in schedules:
            print(f"  [INFO] No schedule found for {cid} — assuming fully available all week")
            schedules[cid] = {
                "schedule_type": "recurring",
                "week_start_date": None,
                "availability": {
                    day: {"AM": True, "MD": True, "PM": True}
                    for day in ["MON", "TUE", "WED", "THU", "FRI", "SAT", "SUN"]
                },
                "notes": "No schedule file — assumed fully available",
            }

    return schedules


def resolve_availability(
    schedules: Dict[str, Dict],
    week_days:  List[date],
) -> Dict[str, Dict[Tuple[date, str], bool]]:
    """
    Expand raw schedule dicts → {collector_id: {(date, tod): bool}}
    for every day in week_days.
    """
    DAY_KEYS = {0: "MON", 1: "TUE", 2: "WED", 3: "THU", 4: "FRI", 5: "SAT", 6: "SUN"}
    result: Dict[str, Dict[Tuple[date, str], bool]] = {}

    for cid, sched in schedules.items():
        avail_map  = sched.get("availability", {})
        sched_type = sched.get("schedule_type", "recurring")
        cid_avail: Dict[Tuple[date, str], bool] = {}

        for d in week_days:
            date_str = d.strftime("%Y-%m-%d")
            day_key  = DAY_KEYS[d.weekday()]

            for tod in TODS:
                if sched_type == "specific" and date_str in avail_map:
                    # Missing TOD within a known date → unavailable (matches xlsx behaviour)
                    cid_avail[(d, tod)] = bool(avail_map[date_str].get(tod, False))
                elif day_key in avail_map:
                    # Missing TOD within a known day → unavailable (matches xlsx behaviour)
                    cid_avail[(d, tod)] = bool(avail_map[day_key].get(tod, False))
                else:
                    # Day not in schedule at all → unavailable (matches xlsx default)
                    cid_avail[(d, tod)] = False

        result[cid] = cid_avail

    return result


def parse_availability_xlsx(
    week_days: List[date],
) -> Dict[str, Dict[Tuple[date, str], bool]]:
    """
    Read Collector_Schedule/Availability.xlsx.

    Layout (each tab = one collector ID):
        row 1:  (blank) (blank) Mon  Tue  Wed  Thu  Fri  Sat  Sun
        row 2:  (blank)  AM     1/blank ...
        row 3:  (blank)  MD     ...
        row 4:  (blank)  PM     ...

    1 = available, blank = not available.
    Returns {collector_id: {(date, tod): bool}} for every day in week_days,
    covering only collectors whose tab appears in the workbook.
    """
    xlsx_path = BASE_DIR / "Availability.xlsx"
    if not xlsx_path.exists():
        # Fallback: check old location inside Collector_Schedule/
        xlsx_path = SCHEDULE_DIR / "Availability.xlsx"
    if not xlsx_path.exists():
        print(f"  [WARN] {xlsx_path.name} not found — skipping xlsx availability")
        return {}

    # Map abbreviated day names (as they appear in the header row) → weekday int
    DAY_NAME_TO_WEEKDAY = {
        "mon": 0, "tue": 1, "wed": 2, "thu": 3,
        "fri": 4, "sat": 5, "sun": 6,
    }
    # Map weekday int → actual date for this week
    weekday_to_date: Dict[int, date] = {d.weekday(): d for d in week_days}

    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    result: Dict[str, Dict[Tuple[date, str], bool]] = {}

    for sheet_name in wb.sheetnames:
        cid = sheet_name.strip().upper()
        if cid not in COLLECTORS:
            print(f"  [WARN] Availability.xlsx tab '{sheet_name}' is not a known collector — skipping")
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue

        # Find the header row: first row that contains a recognized day name
        header_row_idx: Optional[int] = None
        col_to_weekday: Dict[int, int] = {}
        for r_idx, row in enumerate(rows):
            for col_idx, cell_val in enumerate(row):
                if cell_val is None:
                    continue
                key = str(cell_val).strip().lower()
                if key in DAY_NAME_TO_WEEKDAY:
                    col_to_weekday[col_idx] = DAY_NAME_TO_WEEKDAY[key]
                    header_row_idx = r_idx
            if header_row_idx is not None:
                break   # stop after the first row with day names

        if header_row_idx is None:
            print(f"  [WARN] {sheet_name}: no day-name header found — skipping")
            continue

        # Data rows immediately follow the header
        cid_avail: Dict[Tuple[date, str], bool] = {}

        for data_row in rows[header_row_idx + 1:]:
            if not any(v is not None for v in data_row):
                continue
            tod_label = str(data_row[1]).strip().upper() if data_row[1] is not None else ""
            if tod_label not in TODS:
                continue

            for col_idx, weekday in col_to_weekday.items():
                if weekday not in weekday_to_date:
                    continue          # that day-of-week isn't in the current forecast week
                actual_date = weekday_to_date[weekday]
                cell_val = data_row[col_idx] if col_idx < len(data_row) else None
                available = (cell_val == 1)
                cid_avail[(actual_date, tod_label)] = available

        # Fill any missing (date, tod) slots as unavailable (blank = no availability)
        for d in week_days:
            for tod in TODS:
                cid_avail.setdefault((d, tod), False)

        result[cid] = cid_avail
        avail_count = sum(1 for v in cid_avail.values() if v)
        print(f"  {cid}: {avail_count} available slot(s) this week")

    return result


# ─────────────────────────────────────────────────────────────────────────────
# STEP 5 — Parse KML Files
# ─────────────────────────────────────────────────────────────────────────────

def parse_route_coords() -> Dict[str, Tuple[float, float]]:
    """
    Parse all four boro KML files.
    For each named route, compute the centroid of all LineString coordinates.
    Returns {route_code: (lat, lon)}
    """
    boro_kmls = {
        "bronx":     KML_DIR / "aq routes - bronx.kml",
        "brooklyn":  KML_DIR / "aq routes - brooklyn.kml",
        "manhattan": KML_DIR / "aq routes - manhattan.kml",
        "queens":    KML_DIR / "aq routes - queens.kml",
    }
    ns = {"k": "http://www.opengis.net/kml/2.2"}
    coords: Dict[str, Tuple[float, float]] = {}

    for _boro, kml_file in boro_kmls.items():
        try:
            tree = ET.parse(kml_file)
        except (ET.ParseError, OSError) as e:
            print(f"  [WARN] Could not parse {kml_file.name}: {e}")
            continue
        root = tree.getroot()
        for pm in root.findall(".//k:Placemark", ns):
            name_el = pm.find("k:name", ns)
            name    = (name_el.text or "").strip() if name_el is not None else ""
            if name not in KML_NAME_TO_ROUTE:
                continue
            route_code = KML_NAME_TO_ROUTE[name]

            # Gather all coordinate pairs from every LineString
            lats, lons = [], []
            for ls in pm.findall(".//k:LineString/k:coordinates", ns):
                for token in (ls.text or "").split():
                    parts = token.split(",")
                    if len(parts) >= 2:
                        try:
                            lons.append(float(parts[0]))
                            lats.append(float(parts[1]))
                        except ValueError:
                            pass

            if lats:
                coords[route_code] = (
                    sum(lats) / len(lats),
                    sum(lons) / len(lons),
                )

    return coords


def parse_collector_locs() -> Dict[str, Tuple[float, float]]:
    """
    Parse Collector_Locs.kml.
    Returns {collector_id: (lat, lon)}
    """
    kml_file = KML_DIR / "Collector_Locs.kml"
    ns = {"k": "http://www.opengis.net/kml/2.2"}
    locs: Dict[str, Tuple[float, float]] = {}

    try:
        tree = ET.parse(kml_file)
    except (ET.ParseError, OSError) as e:
        print(f"  [WARN] Could not parse {kml_file.name}: {e}")
        return locs
    root = tree.getroot()

    for pm in root.findall(".//k:Placemark", ns):
        name_el = pm.find("k:name", ns)
        name    = (name_el.text or "").strip() if name_el is not None else ""
        if name not in COLLECTOR_KML_NAMES.values():
            continue
        pt = pm.find(".//k:Point/k:coordinates", ns)
        if pt is None:
            continue
        parts = (pt.text or "").strip().split(",")
        if len(parts) >= 2:
            try:
                lon, lat = float(parts[0]), float(parts[1])
                # Reverse lookup: name → collector ID
                for cid, kname in COLLECTOR_KML_NAMES.items():
                    if kname == name:
                        locs[cid] = (lat, lon)
            except ValueError:
                pass

    return locs


# ─────────────────────────────────────────────────────────────────────────────
# STEP 6 — Score Combos
# ─────────────────────────────────────────────────────────────────────────────

def score_combos(
    completions:   Dict[Tuple[str, str, str], int],
    weather:       Dict[Tuple[date, str, str], bool],
    affinity:      Dict[str, Dict[str, int]],
    availability:  Dict[str, Dict[Tuple[date, str], bool]],
    route_coords:  Dict[str, Tuple[float, float]],
    collector_locs: Dict[str, Tuple[float, float]],
) -> List[Dict]:
    """
    Score every (route, tod) combo for CURRENT_SEASON that is still below target.

    Weather dict is keyed by (date, tod, boro) where boro is one of
    BX/MN/QN/BK (per-borough PDF) or "" (city-wide PDF).
    Routes use their own boro's weather; city-wide "" is the fallback.

    Priority order:
      1. PRIMARY   – at least one good-weather day for that TOD this week (filter)
      2. SECONDARY – combos below minimum (< 6) rank above those at/above minimum
                     within each bucket, larger deficit = higher rank
      3. TERTIARY  – has at least one available+comfortable collector
      4. TIEBREAKER – minimum distance from any available collector to route centroid
    """
    # Pre-compute: for each (tod, boro), which dates have good weather?
    # Also build a union across all boros so collector availability (boro-agnostic)
    # can still be pre-filtered.
    good_weather_by_tod_boro: Dict[Tuple[str, str], List[date]] = defaultdict(list)
    good_weather_by_tod_any:  Dict[str, set] = defaultdict(set)
    for (d, tod, boro), is_good in weather.items():
        if is_good:
            good_weather_by_tod_boro[(tod, boro)].append(d)
            good_weather_by_tod_any[tod].add(d)

    def good_days_for(tod: str, boro: str) -> List[date]:
        """Return good-weather days for this tod+boro; fall back to city-wide."""
        specific = good_weather_by_tod_boro.get((tod, boro), [])
        if specific:
            return specific
        return good_weather_by_tod_boro.get((tod, ""), [])

    # Pre-compute: which (collector, tod) combinations have available days?
    # Use the union of good-weather days across all boroughs so we don't miss
    # any day a collector could theoretically be sent out.
    collector_avail_days: Dict[Tuple[str, str], List[date]] = defaultdict(list)
    for cid in COLLECTORS:
        cid_avail = availability.get(cid, {})
        for tod in TODS:
            for d in good_weather_by_tod_any.get(tod, set()):
                if cid_avail.get((d, tod), True):
                    collector_avail_days[(cid, tod)].append(d)

    scored: List[Dict] = []

    for route in ALL_ROUTES:
        boro, neigh = route.split("_")
        for tod in TODS:
            count = completions.get((route, tod, CURRENT_SEASON), 0)

            # Already at target — skip
            if count >= TARGET_COMPLETIONS:
                continue

            # Good-weather days for this TOD, using this route's borough
            gw_days = sorted(good_days_for(tod, boro))
            # PRIMARY filter: must have at least one good-weather day
            if not gw_days:
                continue

            # Collectors available on ≥1 good-weather day for this TOD,
            # excluding those who explicitly rated this route 0 in V2
            regular_collectors = [
                cid for cid in COLLECTORS
                if cid not in LAST_RESORT_COLLECTORS
                and collector_avail_days.get((cid, tod))
                and affinity.get(cid, {}).get(neigh, 1) != 0  # 0 = explicit "do not assign"
            ]
            # Only add last-resort collectors if no regular collectors available
            available_collectors = regular_collectors if regular_collectors else [
                cid for cid in LAST_RESORT_COLLECTORS
                if collector_avail_days.get((cid, tod))
                and affinity.get(cid, {}).get(neigh, 1) != 0
            ]

            # Comfortable = available AND has a positive affinity score (1–3)
            comfortable_collectors = [
                cid for cid in available_collectors
                if affinity.get(cid, {}).get(neigh, 0) > 0
            ]

            # Per-collector affinity score for use in calendar assignment sorting
            affinity_scores: Dict[str, int] = {
                cid: affinity.get(cid, {}).get(neigh, 0)
                for cid in available_collectors
            }

            # SECONDARY: deficit bucket
            below_min = count < MIN_COMPLETIONS
            deficit   = (MIN_COMPLETIONS - count) if below_min else (TARGET_COMPLETIONS - count)

            # TERTIARY: best affinity score among available collectors (higher = better)
            max_affinity_score = max(
                (affinity_scores[cid] for cid in comfortable_collectors),
                default=0
            )

            # TIEBREAKER: minimum home→route transit time among available collectors
            # Uses MTA subway travel time from collector's home station to route's
            # start station, with TOD-specific service patterns (express vs local).
            # Falls back to haversine if transit data unavailable.
            min_dist = 999.0
            if available_collectors:
                for cid in available_collectors:
                    t = _collector_transit_minutes(cid, route, tod=tod)
                    if t < 999.0:
                        min_dist = min(min_dist, t)
                    elif cid in collector_locs:
                        rc = route_coords.get(route)
                        if rc:
                            clat, clon = collector_locs[cid]
                            dist = haversine_km(clat, clon, rc[0], rc[1])
                            min_dist = min(min_dist, dist)

            # Composite sort key (ascending → higher priority = smaller value)
            sort_key = (
                0 if below_min else 1,      # 0 = below minimum → higher priority
                -deficit,                    # larger deficit = more negative = ranks first
                -max_affinity_score,         # higher score = more negative = ranks first
                round(min_dist, 4),          # closer = better
            )

            # Backpack: alternate A/B based on current parity of count
            backpack = "A" if count % 2 == 0 else "B"

            scored.append({
                "route":                  route,
                "boro":                   boro,
                "neigh":                  neigh,
                "tod":                    tod,
                "backpack":               backpack,
                "count":                  count,
                "below_min":              below_min,
                "deficit":                deficit,
                "max_affinity_score":     max_affinity_score,
                "available_collectors":   available_collectors,
                "comfortable_collectors": comfortable_collectors,
                "affinity_scores":        affinity_scores,
                "good_weather_days":      gw_days,
                "min_dist_km":            round(min_dist, 2),
                "sort_key":               sort_key,
            })

    scored.sort(key=lambda x: x["sort_key"])
    return scored


# ─────────────────────────────────────────────────────────────────────────────
# RECALIBRATION DAY  (monthly, both backpacks, prefer bad-weather day)
# ─────────────────────────────────────────────────────────────────────────────

RECAL_TRIGGER_DAYS = 25   # start proposing a recal day (soft — picks worst-weather slot)
RECAL_TARGET_DAYS  = 30   # ideal spacing goal (informational only)
RECAL_MAX_DAYS     = 40   # hard limit — recal is required regardless of field conditions


def parse_last_recal_date() -> Optional[date]:
    """
    Scan Sample_Walks_Log.txt for all RECAL_MM_DD_YYYY entries and
    return the most recent date, or None if no recal has ever been logged.
    """
    try:
        text = WALKS_LOG.read_text(encoding="utf-8", errors="ignore")
    except FileNotFoundError:
        return None
    pattern = re.compile(r"^RECAL_(\d{2})_(\d{2})_(\d{4})$")
    latest: Optional[date] = None
    for raw in text.splitlines():
        m = pattern.match(raw.strip())
        if not m:
            continue
        try:
            d = date(int(m.group(3)), int(m.group(1)), int(m.group(2)))
        except ValueError:
            continue
        if latest is None or d > latest:
            latest = d
    return latest


def recal_status(today: date) -> Tuple[bool, bool, Optional[int]]:
    """
    Returns (propose, required, days_since_last).

    propose   — True when days elapsed >= RECAL_TRIGGER_DAYS (25) or no log entry.
                Scheduler picks the worst-weather slot but data collection takes
                priority; a fully clear week can still proceed without a recal.
    required  — True when days elapsed >= RECAL_MAX_DAYS (40) or no log entry.
                At this point the recal day is hard-blocked in the calendar
                regardless of field conditions.
    days_since — integer day count, or None if no recal on record.
    """
    last = parse_last_recal_date()
    if last is None:
        return True, True, None
    elapsed  = (today - last).days
    propose  = elapsed >= RECAL_TRIGGER_DAYS
    required = elapsed >= RECAL_MAX_DAYS
    return propose, required, elapsed


def pick_recal_day(
    weather:   Dict[Tuple[date, str], bool],
    week_days: List[date],
) -> Optional[date]:
    """
    Choose the best day this week to bring both backpacks to CCNY.

    Primary criterion  — maximise bad-weather TOD count (0-3).
      A fully cloudy day wastes no field capacity; a fully clear day is
      a last resort used only when 25+ days have passed and no better
      week exists.
    Secondary criterion — prefer weekdays (CCNY lab staff availability).
    Tertiary criterion  — earliest date (get it done sooner rather than later
                          so the next 30-day window starts as early as possible).

    Returns None only if week_days is empty.
    """
    if not week_days:
        return None
    ranked = sorted(
        week_days,
        key=lambda d: (
            -sum(1 for tod in TODS if not weather.get((d, tod), False)),  # more bad → first
            0 if d.weekday() < 5 else 1,                                   # weekday beats weekend
            d,                                                              # earlier beats later
        ),
    )
    return ranked[0]


# ─────────────────────────────────────────────────────────────────────────────
# TRANSIT-AWARE HELPERS FOR BACKPACK CLUSTERING & CONTINUITY
# ─────────────────────────────────────────────────────────────────────────────

# Global transit matrix — loaded once in main(), used by clustering & continuity
_TRANSIT_MATRIX: Optional[Dict[str, Dict[str, float]]] = None
# Per-TOD route-to-route matrices  {"AM": {...}, "MD": {...}, "PM": {...}}
_TOD_TRANSIT_MATRICES: Optional[Dict[str, Dict[str, Dict[str, float]]]] = None
# Collector home → route start transit times (minutes)
_COLLECTOR_ROUTE_MATRIX: Optional[Dict[str, Dict[str, float]]] = None
# Per-TOD collector→route matrices {"AM": {...}, "MD": {...}, "PM": {...}}
_TOD_COLLECTOR_ROUTE_MATRICES: Optional[Dict[str, Dict[str, Dict[str, float]]]] = None


def _transit_minutes(route_a: str, route_b: str, tod: Optional[str] = None) -> float:
    """
    Look up transit travel time (minutes) from route_a's end stop to
    route_b's start stop.  If *tod* is given ("AM", "MD", "PM"), uses
    the time-of-day-specific matrix to reflect express/local service
    patterns.  Falls back to 999 if the matrix is not loaded.
    """
    # Try TOD-specific matrix first
    if tod and _TOD_TRANSIT_MATRICES and tod in _TOD_TRANSIT_MATRICES:
        t = _TOD_TRANSIT_MATRICES[tod].get(route_a, {}).get(route_b, None)
        if t is not None:
            return t
    # Fall back to all-day matrix
    if _TRANSIT_MATRIX is None:
        return 999.0
    return _TRANSIT_MATRIX.get(route_a, {}).get(route_b, 999.0)


def _collector_transit_minutes(cid: str, route: str, tod: Optional[str] = None) -> float:
    """
    Look up transit time (minutes) from collector *cid*'s home station
    to *route*'s start stop.  Uses TOD-specific matrix if available.
    """
    # Try TOD-specific matrix first
    if tod and _TOD_COLLECTOR_ROUTE_MATRICES and tod in _TOD_COLLECTOR_ROUTE_MATRICES:
        t = _TOD_COLLECTOR_ROUTE_MATRICES[tod].get(cid, {}).get(route, None)
        if t is not None:
            return t
    # Fall back to all-day matrix
    if _COLLECTOR_ROUTE_MATRIX and cid in _COLLECTOR_ROUTE_MATRIX:
        return _COLLECTOR_ROUTE_MATRIX[cid].get(route, 999.0)
    return 999.0


def _campus_transit_minutes(route: str, campus_bp: str, tod: Optional[str] = None) -> float:
    """
    Transit time (minutes) from *route*'s start stop to a campus.
    *campus_bp* is 'A' (CCNY) or 'B' (LAGCC).  Uses the proxy route
    in the transit matrix (MN_HT for CCNY, QN_LA for LAGCC).
    """
    proxy = CAMPUS_PROXY_ROUTE.get(campus_bp)
    if proxy is None:
        return 999.0
    return _transit_minutes(proxy, route, tod=tod)


def assign_backpacks_by_campus(
    top: List[Dict],
    route_coords: Dict[str, Tuple[float, float]],
) -> Dict[int, str]:
    """
    Assign each top-N route to the backpack whose campus is closer
    by transit time.  A = CCNY (proxy MN_HT), B = LAGCC (proxy QN_LA).

    Falls back to haversine distance from the campus coordinates if
    the transit matrix is not loaded.
    """
    bp_map: Dict[int, str] = {}

    for i, combo in enumerate(top):
        route = combo["route"]
        t_ccny  = _campus_transit_minutes(route, "A")
        t_lagcc = _campus_transit_minutes(route, "B")

        # Haversine fallback when transit data unavailable for this route
        if t_ccny >= 999 and t_lagcc >= 999:
            rc = route_coords.get(route)
            if rc:
                d_ccny  = haversine_km(rc[0], rc[1], CCNY_LAT, CCNY_LON)
                d_lagcc = haversine_km(rc[0], rc[1], LAGCC_LAT, LAGCC_LON)
                bp_map[i] = "A" if d_ccny <= d_lagcc else "B"
            else:
                bp_map[i] = "A" if i % 2 == 0 else "B"
            continue

        bp_map[i] = "A" if t_ccny <= t_lagcc else "B"

    return bp_map


def count_walks_by_collector(season: str = CURRENT_SEASON) -> Dict[str, int]:
    """
    Count completed walks per collector for the given season from Walks_Log.txt.
    Used to load-balance assignments toward collectors with fewer season walks.
    """
    counts: Dict[str, int] = {cid: 0 for cid in COLLECTORS}
    try:
        text = WALKS_LOG.read_text(encoding="utf-8", errors="ignore")
    except FileNotFoundError:
        return counts

    for line in text.splitlines():
        line = line.strip()
        if not line or line.startswith("RECAL"):
            continue
        parts = line.split("_")
        try:
            if len(parts) == 6:                          # new YYYYMMDD format
                col = parts[1]
                raw = parts[4]
                d = date(int(raw[:4]), int(raw[4:6]), int(raw[6:8]))
            elif len(parts) == 9:                        # old MM_DD_YYYY format
                col = parts[1]
                d = date(int(parts[6]), int(parts[4]), int(parts[5]))
            else:
                continue
        except (ValueError, IndexError):
            continue

        if get_season(d) == season and col in counts:
            counts[col] += 1

    return counts


def _continuity_cost(
    cid: str,
    route: str,
    route_coords: Dict[str, Tuple[float, float]],
    collector_weekly_routes: Dict[str, List[str]],
    same_day: bool = False,
    tod: Optional[str] = None,
) -> float:
    """
    Transit-time-weighted continuity cost (minutes) from `route` to
    the routes already assigned to `cid` this week.

    Uses the transit travel time matrix (end-stop of prior route ->
    start-stop of candidate route), with TOD-specific service patterns
    when available.  Same-day sequential assignments are weighted 2x
    because the collector must physically travel between walks that day;
    next-day assignments are weighted 1x.

    Falls back to haversine km if the transit matrix is not available.
    Returns 0.0 if the collector has no prior assignments.
    """
    prior = collector_weekly_routes.get(cid, [])
    if not prior:
        return 0.0

    weight = 2.0 if same_day else 1.0

    if _TRANSIT_MATRIX is not None:
        times = []
        for r in prior:
            t = _transit_minutes(r, route, tod=tod)
            if t < 999:
                times.append(t)
        if times:
            return weight * (sum(times) / len(times))
        # fall through to haversine if no transit data for these routes

    # Haversine fallback (original behaviour)
    if route not in route_coords:
        return 0.0
    rc = route_coords[route]
    dists = []
    for r in prior:
        if r in route_coords:
            pr = route_coords[r]
            dists.append(haversine_km(rc[0], rc[1], pr[0], pr[1]))
    return weight * (sum(dists) / len(dists)) if dists else 0.0


# ─────────────────────────────────────────────────────────────────────────────
# STEP 7a — Print Ranked Table
# ─────────────────────────────────────────────────────────────────────────────

def print_ranked_table(
    scored:       List[Dict],
    route_coords: Dict[str, Tuple[float, float]],
    top_n:        int = 8,
    bp_filter:    Optional[str] = None,
) -> None:
    top = scored[:top_n]
    if bp_filter:
        bp_map = {i: bp_filter for i in range(len(top))}
    else:
        bp_map = assign_backpacks_by_campus(top, route_coords)

    W = 112
    print()
    print("═" * W)
    print("  TOP 8 RECOMMENDED WALKS — RANKED TABLE")
    print("═" * W)
    print(
        f"  {'#':>2}  {'Route':<28}  {'TOD':<4}  "
        f"{'BP':<3}  {'Done':>4}  {'Status':<10}  "
        f"{'Good-Weather Days This Week':<30}  "
        f"Available Collectors  ([comfortable])"
    )
    print("─" * W)

    for rank, (i, combo) in enumerate(zip(range(len(top)), top), 1):
        label      = ROUTE_LABELS.get(combo["route"], combo["route"])
        gw_str     = ", ".join(d.strftime("%a %m/%d") for d in combo["good_weather_days"])
        status_str = "BELOW MIN" if combo["below_min"] else "NEEDS WALK"
        bp         = bp_map[rank - 1]

        coll_parts = []
        for cid in sorted(set(combo["available_collectors"])):
            tag = f"[{cid}]" if cid in combo["comfortable_collectors"] else cid
            coll_parts.append(tag)
        coll_str  = " ".join(coll_parts) if coll_parts else "—"
        dist_note = f"  ({combo['min_dist_km']} km)" if combo["min_dist_km"] < 900 else ""

        print(
            f"  {rank:>2}  {label:<28}  {combo['tod']:<4}  "
            f"{bp:<3}  {combo['count']:>4}  {status_str:<10}  "
            f"{gw_str:<30}  {coll_str}{dist_note}"
        )

    print("═" * W)
    print(
        "  [COLL] = comfortable (route affinity) | BP A = CCNY, BP B = LAGCC (nearest campus) | "
        "Done = Spring completions"
    )
    print()
    return bp_map   # pass through so calendar can reuse the same assignment


# ─────────────────────────────────────────────────────────────────────────────
# SCHEDULE MAP — Folium HTML map with backpack-colored route connections
# ─────────────────────────────────────────────────────────────────────────────

# Backpack colours
BP_COLORS = {"A": "#D82233", "B": "#0062CF"}   # red (A), blue (B)
BP_COLORS_LIGHT = {"A": "#F5A0A8", "B": "#92B8E8"}  # lighter tints for markers


def _generate_schedule_map(
    assignments: List[Dict],
    route_coords: Dict[str, Tuple[float, float]],
    week_start: date,
    week_end: date,
) -> None:
    """
    Create an interactive Folium map showing:
      - Campus pins for CCNY (Backpack A) and LAGCC (Backpack B)
      - Circle markers at each scheduled route centroid
      - Dotted lines connecting routes in chronological order, per backpack
      - Lines and markers colored by backpack (A=red, B=blue)
    Saves as schedule_map.html in BASE_DIR.
    """
    if not assignments:
        print("  No assignments to map.")
        return

    # Centre the map on the mean of all assigned route coords
    lats, lons = [], []
    for e in assignments:
        rc = route_coords.get(e["route"])
        if rc:
            lats.append(rc[0])
            lons.append(rc[1])
    if not lats:
        print("  No geocoded routes to map.")
        return

    m = folium.Map(
        location=[sum(lats) / len(lats), sum(lons) / len(lons)],
        zoom_start=11,
        tiles="CartoDB positron",
    )

    # ── Collector area polylines from borough KMLs ───────────────────────────
    _boro_kml_files = {
        "bronx":     KML_DIR / "aq routes - bronx.kml",
        "brooklyn":  KML_DIR / "aq routes - brooklyn.kml",
        "manhattan": KML_DIR / "aq routes - manhattan.kml",
        "queens":    KML_DIR / "aq routes - queens.kml",
    }
    _boro_route_colors = {"MN": "#a78bfa", "BX": "#fb923c", "BK": "#34d399", "QN": "#60a5fa"}
    _ns_kml = {"k": "http://www.opengis.net/kml/2.2"}
    areas_group = folium.FeatureGroup(name="Collection Areas", show=True)
    for _boro, _kml_file in _boro_kml_files.items():
        try:
            _tree = ET.parse(_kml_file)
        except (ET.ParseError, OSError):
            continue
        for _pm in _tree.getroot().findall(".//k:Placemark", _ns_kml):
            _nm_el = _pm.find("k:name", _ns_kml)
            _nm    = (_nm_el.text or "").strip() if _nm_el is not None else ""
            _rc    = KML_NAME_TO_ROUTE.get(_nm)
            if not _rc:
                continue
            _boro_prefix = _rc.split("_")[0]
            _color = _boro_route_colors.get(_boro_prefix, "#94a3b8")
            for _ls in _pm.findall(".//k:LineString/k:coordinates", _ns_kml):
                _pts = []
                for _tok in (_ls.text or "").split():
                    _p = _tok.split(",")
                    if len(_p) >= 2:
                        try:
                            _pts.append([float(_p[1]), float(_p[0])])
                        except ValueError:
                            pass
                if len(_pts) >= 2:
                    folium.PolyLine(
                        locations=_pts,
                        color=_color,
                        weight=3,
                        opacity=0.55,
                        tooltip=ROUTE_LABELS.get(_rc, _nm),
                    ).add_to(areas_group)
    areas_group.add_to(m)

    # ── Collector home pins ──────────────────────────────────────────────────
    _collector_homes = parse_collector_locs()
    _bp_for_cid = {cid: bp for bp, members in BACKPACK_COLLECTORS.items() for cid in members}
    homes_group = folium.FeatureGroup(name="Collector Homes", show=True)
    for cid, (lat, lon) in _collector_homes.items():
        bp   = _bp_for_cid.get(cid, "A")
        col  = BP_COLORS[bp]
        name = COLLECTOR_ID_TO_NAME.get(cid, cid)
        folium.Marker(
            location=[lat, lon],
            tooltip=f"{name} ({cid}) — BP {bp}",
            popup=folium.Popup(f"<b>{name}</b><br>Backpack {bp}", max_width=160),
            icon=folium.DivIcon(
                html=(
                    f'<div style="background:{col};color:#fff;border:2px solid #fff;'
                    f'border-radius:50%;width:22px;height:22px;display:flex;'
                    f'align-items:center;justify-content:center;font-size:9px;'
                    f'font-weight:700;box-shadow:0 0 4px rgba(0,0,0,.5);">'
                    f'{cid}</div>'
                ),
                icon_size=(22, 22),
                icon_anchor=(11, 11),
            ),
        ).add_to(homes_group)
    homes_group.add_to(m)

    # Group assignments by backpack, sorted chronologically
    for bp in ("A", "B"):
        bp_entries = sorted(
            [e for e in assignments if e["backpack"] == bp],
            key=lambda x: (x["assigned_date"], TODS.index(x["tod"])),
        )
        if not bp_entries:
            continue

        color = BP_COLORS[bp]
        prev_coord: Optional[Tuple[float, float]] = None

        for seq, e in enumerate(bp_entries, 1):
            rc = route_coords.get(e["route"])
            if rc is None:
                continue

            label = ROUTE_LABELS.get(e["route"], e["route"])
            day_str = e["assigned_date"].strftime("%a %m/%d")
            popup_text = (
                f"<b>BP {bp} #{seq}</b><br>"
                f"{label}<br>"
                f"{day_str} {e['tod']}<br>"
                f"Collector: {e['assigned_collector']}"
            )

            # Circle marker at route centroid
            folium.CircleMarker(
                location=[rc[0], rc[1]],
                radius=10,
                color=color,
                fill=True,
                fill_color=color,
                fill_opacity=0.7,
                popup=folium.Popup(popup_text, max_width=220),
                tooltip=f"BP {bp}: {e['route']} ({e['tod']} {day_str})",
            ).add_to(m)

            # Sequence number label
            folium.Marker(
                location=[rc[0], rc[1]],
                icon=folium.DivIcon(
                    html=(
                        f'<div style="font-size:11px; font-weight:bold; '
                        f'color:white; background:{color}; '
                        f'border-radius:50%; width:20px; height:20px; '
                        f'text-align:center; line-height:20px; '
                        f'border:2px solid white; box-shadow:0 0 3px rgba(0,0,0,0.4);">'
                        f'{seq}</div>'
                    ),
                    icon_size=(20, 20),
                    icon_anchor=(10, 10),
                ),
            ).add_to(m)

            # Dotted line from previous route to this one
            if prev_coord is not None:
                folium.PolyLine(
                    locations=[
                        [prev_coord[0], prev_coord[1]],
                        [rc[0], rc[1]],
                    ],
                    color=color,
                    weight=3,
                    opacity=0.75,
                    dash_array="8 6",  # dotted line
                    tooltip=f"BP {bp}: transit link",
                ).add_to(m)

            prev_coord = rc

    # ── Campus markers (CCNY = Backpack A, LAGCC = Backpack B) ──────────────
    campus_markers = [
        (CCNY_LAT, CCNY_LON, "CCNY", "A",
         "CCNY — Marshak Science Building<br>160 Convent Ave<br><i>Backpack A home</i>"),
        (LAGCC_LAT, LAGCC_LON, "LAGCC", "B",
         "LaGuardia Community College<br>31-10 Thomson Ave, LIC<br><i>Backpack B home</i>"),
    ]
    for lat, lon, name, bp, popup_text in campus_markers:
        color = BP_COLORS[bp]
        folium.Marker(
            location=[lat, lon],
            popup=folium.Popup(popup_text, max_width=240),
            tooltip=f"{name} (Backpack {bp})",
            icon=folium.Icon(
                color="red" if bp == "A" else "blue",
                icon="university",
                prefix="fa",
            ),
        ).add_to(m)

    # Legend
    legend_html = """
    <div style="position:fixed; bottom:30px; left:30px; z-index:1000;
                background:white; padding:12px 16px; border-radius:8px;
                border:2px solid #ccc; font-family:Arial,sans-serif;
                font-size:13px; box-shadow:2px 2px 6px rgba(0,0,0,0.2);">
      <b>Schedule Map</b><br>
      <span style="color:#D82233;">&#9679;</span> Backpack A — CCNY &nbsp;
      <span style="color:#0062CF;">&#9679;</span> Backpack B — LAGCC<br>
      <span style="font-size:11px; color:#666;">
        WEEK_RANGE<br>
        🏫 = campus &nbsp; Dotted lines = chronological transit path
      </span>
    </div>
    """.replace(
        "WEEK_RANGE",
        f"{week_start.strftime('%b %d')} - {week_end.strftime('%b %d, %Y')}",
    )
    m.get_root().html.add_child(folium.Element(legend_html))
    folium.LayerControl(collapsed=False).add_to(m)

    map_path = BASE_DIR / "schedule_map.html"
    m.save(str(map_path))
    print(f"  Schedule map saved -> {map_path.name}")


# ─────────────────────────────────────────────────────────────────────────────
# STEP 7b — Build Assignments + Print Two Backpack Calendars
# ─────────────────────────────────────────────────────────────────────────────

def build_weekly_calendar(
    scored:              List[Dict],
    availability:        Dict[str, Dict[Tuple[date, str], bool]],
    weather:             Dict[Tuple[date, str], bool],
    route_coords:        Dict[str, Tuple[float, float]],
    week_start:          date,
    week_end:            date,
    recal_day:           Optional[date] = None,
    top_n:               int = 8,
    season_counts:       Optional[Dict[str, int]] = None,
    bp_filter:           Optional[str] = None,
    preserved_assignments: Optional[List[Dict]] = None,
) -> None:
    """
    Assign each top-N combo to a day + TOD respecting:
      • Good weather on that day+TOD
      • Collector available on that day+TOD
      • No collector scheduled more than once per day
      • Load balancing: prefer collectors with fewer completed walks this season
      • Continuity: among equal-load collectors, prefer the one whose
        already-assigned routes this week are geographically closest
        (greedy nearest-neighbor heuristic)
      • recal_day (if set): that entire day is blocked on BOTH backpack
        calendars for CCNY recalibration — no field walks are placed on it.

    Outputs two separate calendar grids — one for Backpack A, one for B.
    """
    top   = scored[:top_n]
    today = date.today()

    # Backpack assignment: by team (bp_filter) or campus proximity (unfiltered run)
    if bp_filter:
        bp_map = {i: bp_filter for i in range(len(top))}
    else:
        bp_map = assign_backpacks_by_campus(top, route_coords)

    all_week_days = [
        week_start + timedelta(days=i)
        for i in range((week_end - week_start).days + 1)
    ]

    # Separate calendars per backpack
    cal: Dict[str, Dict[date, Dict[str, Optional[Dict]]]] = {
        "A": {d: {tod: None for tod in TODS} for d in all_week_days},
        "B": {d: {tod: None for tod in TODS} for d in all_week_days},
    }

    collector_used_on:     Dict[str, set]        = defaultdict(set)
    collector_week_routes: Dict[str, List[str]]  = defaultdict(list)

    assignments: List[Dict] = []
    unassigned:  List[Dict] = []

    # ── Preserve existing assignments where weather is still good ────────────
    preserved_keys: set = set()   # (route, tod) pairs already locked in
    for a in (preserved_assignments or []):
        try:
            d   = date.fromisoformat(a["date"])
            tod = a["tod"]
            bp  = a["backpack"]
        except (KeyError, ValueError):
            continue
        if d not in cal.get(bp, {}):
            continue                     # day not in this week's grid
        if cal[bp][d][tod] is not None:
            continue                     # slot already taken (e.g. forced CCNY)
        entry = {
            "route":              a["route"],
            "boro":               a.get("boro", a["route"].split("_")[0]),
            "neigh":              a.get("neigh", a["route"].split("_")[-1]),
            "tod":                tod,
            "backpack":           bp,
            "assigned_date":      d,
            "assigned_collector": a["collector"],
            "below_min":          False,
            "count":              0,
            "continuity_min":     0,
            "preserved":          True,
        }
        cal[bp][d][tod] = entry
        collector_used_on[a["collector"]].add(d)
        collector_week_routes[a["collector"]].append(a["route"])
        assignments.append(entry)
        preserved_keys.add((a["route"], tod))
        print(f"  ↩  Preserved: [{bp}] {a['route']} {tod} on {d} → {a['collector']}")

    # Filter top combos — skip any (route, tod) already preserved
    top = [c for c in top if (c["route"], c["tod"]) not in preserved_keys]

    # ── Initialize dynamic walk count tracking ──────────────────────────────────
    # Create a mutable copy of season_counts that will be updated as assignments
    # are made, ensuring load balancing is intelligent across the entire scheduling run
    dynamic_season_counts = (season_counts or {}).copy()

    # ── MRV pre-computation ─────────────────────────────────────────────────
    # For each (day, tod) slot, count how many top combos have that slot as a
    # valid good-weather option.  When choosing which day to place a walk on,
    # prefer the day with the fewest competing combos (most constrained first),
    # so rare slots are preserved for the combos that need them most.
    slot_demand: Dict[Tuple[date, str], int] = defaultdict(int)
    for c in top:
        for d in c["good_weather_days"]:
            slot_demand[(d, c["tod"])] += 1

    # ── Pre-compute constraint levels for intelligent ordering ──────────────────
    # Categorize combos by how constrained they are (fewest eligible collectors = tightest)
    # This ensures hard-to-place assignments are handled first while flexibility exists
    combo_constraints: List[Tuple[int, int, Dict]] = []
    for combo_idx, combo in enumerate(top):
        bp = bp_map[combo_idx]
        tod = combo["tod"]
        # Count eligible collectors for this combo across all good-weather days
        eligible_count = 0
        bp_team = BACKPACK_COLLECTORS.get(bp, set())
        for cid in combo["available_collectors"]:
            if cid not in bp_team:
                continue
            # Count how many days this collector is available and not yet used that day
            available_days = sum(
                1 for d in combo["good_weather_days"]
                if d > today and d in cal[bp] and
                   availability.get(cid, {}).get((d, tod), True) and
                   (recal_day is None or d != recal_day) and
                   cal[bp][d][tod] is None and
                   sum(1 for t in TODS if cal[bp][d].get(t) is not None and cal[bp][d][t].get("assigned_collector") == cid) < 2
            )
            if available_days > 0:
                eligible_count += 1
        # Sort by: (constraint severity (fewest options first), original index)
        # Constraint 0 = unplaceable, 1 = forced (1 option), 2 = constrained (2-3), 3 = flexible (4+)
        if eligible_count == 0:
            constraint_level = 0
        elif eligible_count == 1:
            constraint_level = 1  # forced
        elif eligible_count <= 3:
            constraint_level = 2  # constrained
        else:
            constraint_level = 3  # flexible
        combo_constraints.append((constraint_level, combo_idx, combo))

    # Sort by constraint level (ascending), then by original index for stability
    # This processes tight constraints first: forced → constrained → flexible
    combo_constraints.sort(key=lambda x: (x[0], x[1]))

    # Reorder top based on constraint level (but keep unplaceable at end)
    top_sorted = [combo for _, _, combo in combo_constraints if _ > 0]
    top_unplaceable = [combo for _, _, combo in combo_constraints if _ == 0]
    top = top_sorted + top_unplaceable

    # Update bp_map indices to match new ordering
    new_bp_map = {}
    for new_idx, (_, old_idx, _) in enumerate(combo_constraints):
        new_bp_map[new_idx] = bp_map[old_idx]
    bp_map = new_bp_map

    for idx, combo in enumerate(top):
        tod = combo["tod"]
        bp  = bp_map[idx]
        placed = False

        # MRV: least-demanded strictly-future days first
        gw_days_sorted = sorted(
            (d for d in combo["good_weather_days"] if d > today),
            key=lambda d: (slot_demand[(d, tod)], d),
        )

        for d in gw_days_sorted:
            if d not in cal[bp]:
                continue
            if d <= today:
                continue  # never schedule today or any past day
            if recal_day is not None and d == recal_day:
                continue  # entire day reserved for CCNY recalibration
            if cal[bp][d][tod] is not None:
                continue  # slot taken by this backpack

            # Build scored candidate list:
            #   Intelligent scoring: prioritize affinity first, then continuity, then load balance
            #   This prevents early greedy over-assignment to low-count collectors
            bp_team = BACKPACK_COLLECTORS.get(bp, set())
            eligible = []
            for cid in combo["comfortable_collectors"] + [
                c for c in combo["available_collectors"]
                if c not in combo["comfortable_collectors"]
            ]:
                if cid not in bp_team:
                    continue  # enforce: BP-A → CCNY only, BP-B → LaGCC only
                if not availability.get(cid, {}).get((d, tod), True):
                    continue
                # Count walks already assigned to this collector on day d (allow up to 2)
                walks_today = sum(1 for tod_check in TODS if cal[bp][d].get(tod_check) is not None and cal[bp][d][tod_check].get("assigned_collector") == cid)
                if walks_today >= 2:
                    continue
                # affinity_penalty: lower = preferred (score 3 → 0, score 0 → 3, unrated → 3)
                affinity_penalty = 3 - combo["affinity_scores"].get(cid, 0)
                # Use dynamic walk count updated in real-time as assignments are made
                season_walks     = dynamic_season_counts.get(cid, 0)
                # same_day=True when collector has OTHER walk on this same day OR previous day
                # (consecutive-day AND intra-day sequencing — transit ease matters more)
                prev_day = d - timedelta(days=1)
                walks_today_check = sum(1 for tod_check in TODS if cal[bp][d].get(tod_check) is not None and cal[bp][d][tod_check].get("assigned_collector") == cid)
                is_consecutive = walks_today_check > 0 or prev_day in collector_used_on.get(cid, set())
                cont_cost        = _continuity_cost(
                    cid, combo["route"], route_coords, collector_week_routes,
                    same_day=is_consecutive, tod=tod,
                )
                # Priority: affinity first -> continuity second -> load balance third -> alpha
                # This ensures good matches are preserved, transit continuity is maintained,
                # and load balancing happens as a tertiary tiebreaker
                eligible.append((affinity_penalty, cont_cost, season_walks, cid))

            if not eligible:
                continue

            eligible.sort(key=lambda x: (x[0], x[1], x[2], x[3]))
            _affinity, _cont_min, _sea_walks, chosen = eligible[0]

            entry = {
                **combo,
                "assigned_date":      d,
                "assigned_collector": chosen,
                "backpack":           bp,
                "continuity_min":     round(_cont_min, 1),
                "season_walks":       _sea_walks,
            }
            cal[bp][d][tod] = entry
            collector_used_on[chosen].add(d)
            collector_week_routes[chosen].append(combo["route"])
            assignments.append(entry)
            # ── Update dynamic walk count for intelligent load balancing ──────────
            # Increment the collector's count so subsequent assignments see the updated value
            # This prevents the algorithm from repeatedly assigning to the same person
            dynamic_season_counts[chosen] = dynamic_season_counts.get(chosen, 0) + 1
            placed = True
            break

        if not placed:
            if not combo.get("available_collectors"):
                _reason = "no_available_collectors"
            else:
                _reason = "no_open_slot"
            unassigned.append({**combo, "backpack": bp, "reason": _reason})

    # ── Helper: print one backpack's calendar ───────────────────────────────
    def _print_cal(bp_label: str) -> None:
        W   = 112
        col = 15
        print()
        print("═" * W)
        recal_note = (
            f"  —  ★ RECAL: {recal_day.strftime('%a %m/%d')}"
            if recal_day else ""
        )
        campus_name = CAMPUS_COORDS[bp_label][2]  # "CCNY" or "LAGCC"
        print(
            f"  BACKPACK {bp_label} ({campus_name}) — WEEKLY CALENDAR  "
            f"({week_start.strftime('%b %d')} – {week_end.strftime('%b %d, %Y')})"
            f"{recal_note}"
        )
        print("═" * W)

        # Mark recal column header with ★
        day_labels = []
        for d in all_week_days:
            lbl = d.strftime("%a %m/%d")
            day_labels.append(f"★{lbl}" if recal_day and d == recal_day else lbl)

        print("  TOD  |" + "".join(f"  {lbl:<{col}}" for lbl in day_labels))
        print("─" * W)

        for tod in TODS:
            cells = []
            for d in all_week_days:
                marker = "*" if d < today else " "
                if recal_day and d == recal_day:
                    cell = "★ RECAL CCNY"
                    cells.append(f"{marker}{cell:<{col - 1}}")
                    continue
                entry   = cal[bp_label][d][tod]
                wx_good = weather.get((d, tod), False)
                if entry:
                    cell = f"{entry['route']}({entry['assigned_collector']})"
                elif not wx_good:
                    cell = "☁ BAD WX"
                else:
                    cell = "· open"
                cells.append(f"{marker}{cell:<{col - 1}}")
            print(f"  {tod:<4} |" + "".join(f"  {c}" for c in cells))

        print("─" * W)
        print(
            f"  * = past   ☁ = cloudy TOD   (ID) = assigned collector"
            + (f"   ★ = CCNY recalibration day (both backpacks)" if recal_day else "")
        )
        print()

        # Per-day detail for this backpack
        bp_entries = sorted(
            [e for e in assignments if e["backpack"] == bp_label],
            key=lambda x: (x["assigned_date"], TODS.index(x["tod"])),
        )
        if not bp_entries:
            print(f"  No walks assigned to Backpack {bp_label} this week.")
            print()
            return

        print(f"  {'Date':<14} {'TOD':<5} {'Route':<28} {'Collector':<8} "
              f"{'Done':<5} {'Status':<10} {'Transit Continuity'}")
        print("  " + "─" * 84)
        for e in bp_entries:
            d_     = e["assigned_date"]
            label_ = ROUTE_LABELS.get(e["route"], e["route"])
            status = "BELOW MIN" if e["below_min"] else "NEEDS WALK"
            past   = " *" if d_ < today else ""
            cont   = f"{e['continuity_min']} min transit from prior" if e.get("continuity_min", 0) > 0 else "first walk"
            print(
                f"  {d_.strftime('%a %b %d'):<14} {e['tod']:<5} {label_:<28} "
                f"{e['assigned_collector']:<8} {e['count']:<5} {status:<10} {cont}{past}"
            )
        print()

    if bp_filter:
        _print_cal(bp_filter)
    else:
        _print_cal("A")
        _print_cal("B")

    # ── Unassigned ───────────────────────────────────────────────────────────
    if unassigned:
        W = 112
        print(f"  ⚠  {len(unassigned)} combo(s) could not be placed in any open slot:")
        for combo in unassigned:
            label = ROUTE_LABELS.get(combo["route"], combo["route"])
            gw    = [d.strftime("%a %m/%d") for d in combo["good_weather_days"]]
            print(
                f"     • [BP {combo['backpack']}] {label} [{combo['tod']}]  "
                f"good days: {gw}  available: {combo['available_collectors']}"
            )
        print()
        print("═" * W)
        print()

    # ── Schedule map ──────────────────────────────────────────────────────────
    _generate_schedule_map(assignments, route_coords, week_start, week_end)

    # ── JSON export for dashboard ─────────────────────────────────────────────
    schedule_data = {
        "generated":    str(date.today()),
        "generated_at": datetime.now().isoformat(),
        "week_start":   str(week_start),
        "week_end":     str(week_end),
        "recal_day":    str(recal_day) if recal_day else None,
        "assignments": [
            {
                "route":     e["route"],
                "label":     ROUTE_LABELS.get(e["route"], e["route"]),
                "boro":      e["boro"],
                "neigh":     e["neigh"],
                "tod":       e["tod"],
                "backpack":  e["backpack"],
                "collector": e["assigned_collector"],
                "date":      str(e["assigned_date"]),
                "preserved": e.get("preserved", False),
            }
            for e in assignments
        ],
        "unassigned": [
            {
                "route":   e["route"],
                "label":   ROUTE_LABELS.get(e["route"], e["route"]),
                "tod":     e["tod"],
                "backpack": e["backpack"],
                "reason":  e.get("reason", "unknown"),
            }
            for e in unassigned
        ],
    }
    out_path = Path(__file__).parent / "schedule_output.json"
    with open(out_path, "w") as f:
        json.dump(schedule_data, f, indent=2)
    print(f"  Schedule saved -> schedule_output.json")


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(add_help=False)
    parser.add_argument("--backpack", choices=["A", "B"], default=None,
                        help="Limit output to one backpack (A=CCNY, B=LaGCC). Default: both.")
    args, _ = parser.parse_known_args()
    bp_filter = args.backpack

    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        print("[ERROR] ANTHROPIC_API_KEY environment variable is not set.")
        print("        Set it before running:  export ANTHROPIC_API_KEY=sk-ant-...")
        sys.exit(1)

    client = anthropic.Anthropic(api_key=api_key)

    print()
    print("╔══════════════════════════════════════════════════╗")
    print("║   NYC Walk Scheduler — Field Data Campaign       ║")
    print(f"║   Run date: {date.today().strftime('%B %d, %Y'):<38}║")
    if bp_filter:
        campus = "CCNY" if bp_filter == "A" else "LaGCC"
        print(f"║   Backpack filter: {bp_filter} ({campus}){' ' * (28 - len(campus))}║")
    print("╚══════════════════════════════════════════════════╝")
    print()

    # ── Step 0 — Load transit matrix ─────────────────────────────────────
    global _TRANSIT_MATRIX, _COLLECTOR_ROUTE_MATRIX
    global _TOD_TRANSIT_MATRICES, _TOD_COLLECTOR_ROUTE_MATRICES
    transit_json = BASE_DIR / "transit_matrix.json"
    if transit_json.exists():
        try:
            with open(transit_json, encoding="utf-8") as f:
                _tm = json.load(f)
        except (json.JSONDecodeError, OSError) as e:
            print(f"  [WARN] Could not load transit_matrix.json: {e} — using haversine fallback")
            _tm = {}
        _TRANSIT_MATRIX = _tm.get("route_to_route_minutes")
        _COLLECTOR_ROUTE_MATRIX = _tm.get("collector_to_route_minutes")
        _TOD_TRANSIT_MATRICES = _tm.get("tod_route_to_route_minutes")
        _TOD_COLLECTOR_ROUTE_MATRICES = _tm.get("tod_collector_to_route_minutes")
        n_routes = len(_TRANSIT_MATRIX) if _TRANSIT_MATRIX else 0
        n_collectors = len(_COLLECTOR_ROUTE_MATRIX) if _COLLECTOR_ROUTE_MATRIX else 0
        tod_info = ""
        if _TOD_TRANSIT_MATRICES:
            tod_info = f", TOD variants: {', '.join(sorted(_TOD_TRANSIT_MATRICES.keys()))}"
        print(f"  Transit matrix loaded: {n_routes} routes, {n_collectors} collector homes{tod_info}\n")
    else:
        print("  transit_matrix.json not found — using haversine fallback")
        print("  (run `python transit_matrix.py` to generate it)\n")

    # ── Step 1 ──────────────────────────────────────────────────────────────
    print("▶ Step 1  Parsing walk log …")
    completions = parse_walks_log()
    total = sum(completions.values())
    print(f"  {total} walks logged across {len(completions)} unique route+TOD+season combos\n")

    # ── Step 2 ──────────────────────────────────────────────────────────────
    print("▶ Step 2  Loading current-week forecast …")
    forecast_pdf, week_start, week_end = find_current_week_forecast()
    weather = parse_forecast_pdf(forecast_pdf)
    good = sum(1 for v in weather.values() if v)
    boros_in_wx = sorted({b for (_, _, b) in weather if b})
    wx_scope = f"per-borough ({', '.join(boros_in_wx)})" if boros_in_wx else "city-wide"
    print(f"  {good}/{len(weather)} day+TOD slots have good weather (≤{CLOUD_THRESHOLD}% cloud, {wx_scope})\n")

    # ── Load existing schedule — preserve assignments still in good weather ──
    preserved_assignments: List[Dict] = []
    existing_path = Path(__file__).parent / "schedule_output.json"
    confirmations_path = Path(__file__).parent / "schedule_confirmations.json"
    # Load confirmation states so denied slots get re-scheduled
    _confirmations: dict = {}
    if confirmations_path.exists():
        try:
            _confirmations = json.loads(confirmations_path.read_text(encoding="utf-8"))
        except Exception:
            pass
    if existing_path.exists():
        try:
            with open(existing_path, encoding="utf-8") as _f:
                _existing = json.load(_f)
            for _a in _existing.get("assignments", []):
                _d   = date.fromisoformat(_a["date"])
                _tod = _a["tod"]
                # Build the canonical assignment ID used by the confirmation system
                _aid = f"{_a['route']}_{_tod}_{_a['date']}"
                _conf_status = _confirmations.get(_aid, {}).get("status", "pending")
                # Skip if explicitly denied — it will be re-scheduled
                if _conf_status == "denied":
                    print(f"  ✗  Denied by scheduler — re-queuing: {_a['route']} {_tod} {_a['date']}")
                    continue
                # Skip if outside the current week
                if not (week_start <= _d <= week_end):
                    continue
                # Hard-frozen: confirmed by a scheduler OR preserved=True flag — keep regardless of weather
                if _conf_status == "confirmed" or _a.get("preserved", False):
                    preserved_assignments.append(_a)
                    print(f"  🔒 Frozen:    {_a['route']} {_tod} {_a['date']} → {_a['collector']}")
                # Otherwise keep only if weather still good in new forecast
                elif weather.get((_d, _tod), False):
                    preserved_assignments.append(_a)
            if preserved_assignments:
                print(f"  ↩  {len(preserved_assignments)} assignment(s) preserved "
                      f"(frozen or still-good-weather)\n")
        except Exception as _e:
            print(f"  [Note] Could not read existing schedule for preservation: {_e}\n")

    # Week days
    week_days = [
        week_start + timedelta(days=i)
        for i in range((week_end - week_start).days + 1)
    ]

    # ── Step 3 ──────────────────────────────────────────────────────────────
    print("▶ Step 3  Parsing preferred routes …")
    affinity = parse_preferred_routes()
    collectors_with_prefs = sum(1 for v in affinity.values() if any(s > 0 for s in v.values()))
    print(f"  {collectors_with_prefs} collectors have route preferences\n")

    # ── Step 4 ──────────────────────────────────────────────────────────────
    print("▶ Step 4  Parsing collector availability …")

    # Primary: structured xlsx (fast, no API cost)
    print("  Reading Availability.xlsx …")
    availability = parse_availability_xlsx(week_days)

    # Fallback: vision/OCR for any collector not covered by the xlsx
    missing = [c for c in COLLECTORS if c not in availability]
    if missing:
        print(f"  {len(missing)} collector(s) not in xlsx — falling back to vision: {missing}")
        if SCHEDULE_DIR.is_dir():
            schedules    = parse_collector_schedules(client)
            vision_avail = resolve_availability(schedules, week_days)
        else:
            print(f"  [WARN] {SCHEDULE_DIR.name}/ not found — skipping vision fallback")
            schedules, vision_avail = {}, {}
        for cid in missing:
            if cid in vision_avail:
                availability[cid] = vision_avail[cid]
                print(f"    {cid}: loaded via vision")
            else:
                # Last resort: assume fully unavailable (blank xlsx = no availability)
                availability[cid] = {(d, tod): False for d in week_days for tod in TODS}
                print(f"    {cid}: no data found — marked unavailable")
    else:
        print("  All collectors covered by xlsx — skipping vision parsing")
    print()

    # ── Override: TER unavailable Monday PM on the 4th Monday of each month ──
    for d in week_days:
        if d.weekday() == 0 and (d.day - 1) // 7 == 3:  # 4th Monday (days 22–28)
            if "TER" in availability:
                availability["TER"][(d, "PM")] = False
                print(f"  [Override] TER Monday PM blocked (4th Monday): {d}")

    # ── Backpack filter: restrict collectors to the active team ─────────────
    if bp_filter:
        allowed = set(BACKPACK_COLLECTORS[bp_filter])
        if bp_filter == LAST_RESORT_BACKPACK:
            allowed.update(LAST_RESORT_COLLECTORS)
        campus = "CCNY" if bp_filter == "A" else "LaGCC"
        print(f"  [BP-{bp_filter} / {campus}] Restricting to collectors: {sorted(allowed)}\n")
        for cid in availability:
            if cid not in allowed:
                availability[cid] = {k: False for k in availability[cid]}

    # ── Step 5 ──────────────────────────────────────────────────────────────
    print("▶ Step 5  Parsing route and collector coordinates …")
    route_coords   = parse_route_coords()
    collector_locs = parse_collector_locs()
    print(f"  {len(route_coords)} routes geocoded, {len(collector_locs)} collectors located\n")

    # ── Step 6 ──────────────────────────────────────────────────────────────
    print("▶ Step 6  Scoring combos …")
    scored = score_combos(
        completions, weather, affinity,
        availability, route_coords, collector_locs,
    )
    print(f"  {len(scored)} scoreable combos (good weather + below target)\n")

    # ── Step 7a ─────────────────────────────────────────────────────────────
    print("▶ Step 7a Ranked recommendations …")
    print_ranked_table(scored, route_coords, top_n=20, bp_filter=bp_filter)

    # ── Recalibration day ───────────────────────────────────────────────────
    today_                       = date.today()
    propose, required, days_since = recal_status(today_)
    recal_day: Optional[date]    = None

    if not propose:
        days_until = RECAL_TRIGGER_DAYS - days_since
        print(
            f"  ✓ Last recalibration was {days_since} day(s) ago — "
            f"next recal window opens in ~{days_until} day(s) "
            f"(at {RECAL_TRIGGER_DAYS}-day mark, hard limit {RECAL_MAX_DAYS} days).\n"
        )
    else:
        recal_day = pick_recal_day(weather, week_days)
        bad_count = sum(1 for tod in TODS if not weather.get((recal_day, tod), False))

        if days_since is None:
            status_tag = "No recalibration on record — scheduling first recal."
        elif required:
            status_tag = (
                f"⚠  REQUIRED — {days_since} days since last recal "
                f"(hard limit: {RECAL_MAX_DAYS} days). "
                f"Calendar day is blocked regardless of field conditions."
            )
        else:
            days_left = RECAL_MAX_DAYS - days_since
            status_tag = (
                f"{days_since} days since last recal — recal window open. "
                f"Proposing worst-weather slot; hard limit in ~{days_left} day(s)."
            )

        print(
            f"  ★ {status_tag}\n"
            f"    Proposed recal day : {recal_day.strftime('%A %b %d')}  "
            f"({bad_count}/3 TODs have bad weather)\n"
            f"    → When complete, append  RECAL_{recal_day.strftime('%m_%d_%Y')}  "
            f"to Sample_Walks_Log.txt\n"
        )

        # Only hard-block the calendar when required (40+ days)
        if not required:
            recal_day = None   # proposed day shown in message; calendar stays open

    # ── Step 7b ─────────────────────────────────────────────────────────────
    print("▶ Step 7b Weekly calendars (one per backpack) …")
    season_counts = count_walks_by_collector()
    low, high = min(season_counts.values()), max(season_counts.values())
    print(f"  Season walk load: min={low}, max={high} — "
          f"balancing toward collectors with fewer walks\n")
    build_weekly_calendar(
        scored, availability, weather, route_coords,
        week_start, week_end,
        recal_day=recal_day,
        top_n=30,
        season_counts=season_counts,
        bp_filter=bp_filter,
        preserved_assignments=preserved_assignments,
    )


if __name__ == "__main__":
    main()
